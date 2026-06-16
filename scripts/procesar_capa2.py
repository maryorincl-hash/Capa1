#!/usr/bin/env python3
"""
CAPA 2 — Cliente agenda cita pero NO compra
Transcribe MP3s y analiza con Claude para identificar señales de riesgo
comercial durante la llamada que predijeron la no compra en la cita.

Carpeta de entrada:  LeadsCompra\llamadas_capa2\  (colocar los MP3 aquí)
Salida Excel:        LeadsCompra\datos\Analisis_Capa2_No_Compro.xlsx

Antes de correr:
  $env:ANTHROPIC_API_KEY = (Get-ItemProperty "HKCU:\Environment" -Name "ANTHROPIC_API_KEY").ANTHROPIC_API_KEY
  $ffmpegPath = (Get-ChildItem "$env:LOCALAPPDATA\Microsoft\WinGet\Packages\" -Recurse -Filter "ffmpeg.exe").DirectoryName
  $env:PATH = "$ffmpegPath;$env:PATH"
  python scripts\procesar_capa2.py
"""

import os, json, re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import anthropic
import whisper

# ── configuración ─────────────────────────────────────────────────────────────
LLAMADAS_DIR  = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\llamadas_capa2")
DATOS_DIR     = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos")
EXCEL_SALIDA  = DATOS_DIR / "Analisis_Capa2_No_Compro.xlsx"
PROGRESO_FILE = DATOS_DIR / ".progreso_capa2.json"

MODELO_WHISPER = "small"
MODELO_CLAUDE  = "claude-haiku-4-5"

# ── columnas ──────────────────────────────────────────────────────────────────
COLUMNAS_BASE = [
    "Nombre de Archivo",
    "Cliente",
    "Contacto Efectivo",
    "Tiempo de Conversación",
    "Actitud Inicial",
    "Vehículo / Año",
    "Patente",
    "Licencia Clase B",
    "Sin Encargo Robo",
    "Sin Remate",
    ">30 Días Titular",
    "Prenda / Financiera",
    "Gestión de Objeciones / Dudas",
    "Incentivo / Gancho Comercial",
    "Estado / Agendamiento",
]
COLUMNAS_ANALISIS = [
    "Transcripción",
    "¿Qué generó el agendamiento?",
    "Expectativa de precio del cliente",
    "Señales de duda durante la llamada",
    "Brecha precio-expectativa",
    "Riesgo de no show",
    "Riesgo comercial identificado",
    "Condiciones especiales mencionadas",
    "Recomendación para el inspector",
    "Perfil del cliente para el inspector",
]
TODAS_COLUMNAS = COLUMNAS_BASE + COLUMNAS_ANALISIS

# ── colores de cabecera ───────────────────────────────────────────────────────
COLOR_BASE     = "D97706"  # amarillo-naranja — agendaron pero no compraron
COLOR_ANALISIS = "78350F"  # marron — analisis comercial

# ── prompt Claude ─────────────────────────────────────────────────────────────
PROMPT = """Eres un analista senior de ventas de autos usados en Chile.
Esta llamada TERMINO EN CITA AGENDADA (el cliente acepto la inspeccion), pero la visita NO resulto en compra.
Tu tarea es analizar que senales durante la llamada pudieron haber predicho ese resultado, y que informacion necesita el inspector para prepararse.

TRANSCRIPCION:
{transcripcion}

Responde UNICAMENTE con JSON valido, sin texto adicional:
{{
    "cliente": "nombre completo del cliente o Desconocido",
    "contacto_efectivo": "Si",
    "tiempo_conversacion": "duracion estimada MM:SS min segun extension del dialogo",
    "actitud_inicial": "Receptivo / Inseguro / Desconfiado / Esceptico / Apurado / Molesto / Tranquilo",
    "vehiculo_anio": "Marca Modelo (AAAA) o No mencionado",
    "patente": "patente chilena si se menciona, o No mencionada",
    "licencia_clase_b": "Si / No / No menciona",
    "sin_encargo_robo": "Si / No / No menciona",
    "sin_remate": "Si / No / No menciona",
    "dias_titular": "Si / No / No menciona",
    "prenda_financiera": "Si (incluir institucion si se nombra) / No / No menciona",
    "gestion_objeciones": "descripcion del manejo de objeciones durante la llamada. Max 100 palabras.",
    "incentivo_gancho": "gancho o argumento que uso el agente para cerrar la cita. Max 60 palabras.",
    "estado_agendamiento": "Agendado: [dia hora (sucursal/ciudad)] segun lo acordado",
    "factor_agendamiento": "que fue lo determinante para que el cliente aceptara agendar: argumento del agente, urgencia del cliente, incentivo, etc.",
    "expectativa_precio": "precio o rango que el cliente menciono esperar recibir por su auto, o 'No menciono precio esperado'",
    "senales_duda": "frases, dudas o reservas que el cliente expreso durante la llamada que sugieren que no estaba plenamente convencido. Citar textual si es posible. 'Sin señales evidentes' si no hubo.",
    "brecha_precio_expectativa": "Alta — el cliente espera mucho mas de lo razonable / Media / Baja — expectativa alineada / No se menciona precio",
    "riesgo_no_show": "Alto / Medio / Bajo — probabilidad de que el cliente no se presente a la cita basado en la llamada",
    "riesgo_comercial": "principal riesgo comercial identificado en la llamada para la cita: precio, condicion del auto, desconfianza, prenda, etc.",
    "condiciones_especiales": "condiciones, montos minimos, plazos o exigencias especiales que el cliente menciono. 'Ninguna' si no hubo.",
    "recomendacion_inspector": "recomendacion especifica y concreta para el inspector antes de la cita basada en lo que dijo el cliente. Max 100 palabras.",
    "perfil_inspector": "resumen ejecutivo para el inspector: quien es el cliente, que auto tiene, que precio espera, cuales son los riesgos, que necesita para cerrar. Max 150 palabras."
}}"""


# ── utilidades ────────────────────────────────────────────────────────────────

def cargar_progreso():
    if PROGRESO_FILE.exists():
        with open(PROGRESO_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}

def guardar_progreso(prog):
    PROGRESO_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(PROGRESO_FILE, "w", encoding="utf-8") as f:
        json.dump(prog, f, ensure_ascii=False, indent=2)

def transcribir(modelo_whisper, mp3_path):
    try:
        resultado = modelo_whisper.transcribe(str(mp3_path), language="es", fp16=False)
        return resultado["text"].strip()
    except Exception as e:
        return f"Error transcripcion: {e}"

def analizar_con_claude(client, transcripcion, archivo):
    texto = transcripcion[:9000]
    try:
        response = client.messages.create(
            model=MODELO_CLAUDE,
            max_tokens=1300,
            messages=[{"role": "user", "content": PROMPT.format(transcripcion=texto)}]
        )
        raw = response.content[0].text.strip()
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        return json.loads(raw)
    except Exception as e:
        print(f"  ERROR Claude {archivo}: {e}")
        return None

def crear_excel_vacio():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analisis Capa 2"
    for j, col in enumerate(TODAS_COLUMNAS, 1):
        cell = ws.cell(row=1, column=j, value=col)
        color = COLOR_BASE if j <= len(COLUMNAS_BASE) else COLOR_ANALISIS
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    return wb, ws

def escribir_fila(ws, fila_num, datos_dict):
    for j, col in enumerate(TODAS_COLUMNAS, 1):
        cell = ws.cell(row=fila_num, column=j, value=datos_dict.get(col, ""))
        cell.alignment = Alignment(wrap_text=True, vertical="top")

def ajustar_anchos(ws):
    ANCHOS = {
        "Nombre de Archivo": 22, "Cliente": 22, "Transcripcion": 60,
        "Gestion de Objeciones / Dudas": 40,
        "Perfil del cliente para el inspector": 50,
        "Recomendacion para el inspector": 42,
        "Senales de duda durante la llamada": 42,
        "Factor que genero el agendamiento": 38,
    }
    for col in ws.columns:
        header = col[0].value or ""
        max_len = max((min(len(str(c.value or "").split("\n")[0]), 60) for c in col[1:]), default=0)
        ws.column_dimensions[col[0].column_letter].width = ANCHOS.get(header, max(14, max_len + 2))

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  CAPA 2 - CLIENTE AGENDA CITA PERO NO COMPRA")
    print("=" * 60)

    LLAMADAS_DIR.mkdir(parents=True, exist_ok=True)
    DATOS_DIR.mkdir(parents=True, exist_ok=True)

    mp3s = sorted(LLAMADAS_DIR.glob("*.mp3"))
    if not mp3s:
        print(f"\n  No hay MP3 en: {LLAMADAS_DIR}")
        print("  Coloca alli los archivos de llamadas con cita agendada sin compra y vuelve a correr.")
        return

    print(f"\n  MP3s encontrados: {len(mp3s)}")
    print(f"  Salida: {EXCEL_SALIDA}\n")

    progreso = cargar_progreso()
    ya_procesados = set(progreso.keys())
    pendientes = [p for p in mp3s if p.name not in ya_procesados]

    if not pendientes:
        print("  Todos los archivos ya estaban procesados. Regenerando Excel...")
    else:
        print(f"  Pendientes: {len(pendientes)} (ya procesados: {len(ya_procesados)})")

    print("\n  Cargando modelo Whisper...")
    modelo_whisper = whisper.load_model(MODELO_WHISPER)
    client = anthropic.Anthropic()

    for idx, mp3 in enumerate(pendientes, 1):
        print(f"  [{idx}/{len(pendientes)}] {mp3.name}", end=" -> ", flush=True)

        print("Transcribiendo...", end=" ", flush=True)
        transcripcion = transcribir(modelo_whisper, mp3)

        if transcripcion.startswith("Error"):
            print(f"FALLO: {transcripcion}")
            progreso[mp3.name] = {"Transcripcion": transcripcion}
            continue

        print("Analizando...", end=" ", flush=True)
        datos = analizar_con_claude(client, transcripcion, mp3.name)
        if datos is None:
            progreso[mp3.name] = {"Transcripcion": transcripcion}
            continue

        progreso[mp3.name] = {
            "Nombre de Archivo":                          mp3.name,
            "Cliente":                                    datos.get("cliente", "Desconocido"),
            "Contacto Efectivo":                          datos.get("contacto_efectivo", "Si"),
            "Tiempo de Conversacion":                     datos.get("tiempo_conversacion", "N/A"),
            "Actitud Inicial":                            datos.get("actitud_inicial", ""),
            "Vehiculo / Anio":                            datos.get("vehiculo_anio", "No mencionado"),
            "Patente":                                    datos.get("patente", "No mencionada"),
            "Licencia Clase B":                           datos.get("licencia_clase_b", "No menciona"),
            "Sin Encargo Robo":                           datos.get("sin_encargo_robo", "No menciona"),
            "Sin Remate":                                 datos.get("sin_remate", "No menciona"),
            ">30 Dias Titular":                           datos.get("dias_titular", "No menciona"),
            "Prenda / Financiera":                        datos.get("prenda_financiera", "No menciona"),
            "Gestion de Objeciones / Dudas":              datos.get("gestion_objeciones", ""),
            "Incentivo / Gancho Comercial":               datos.get("incentivo_gancho", ""),
            "Estado / Agendamiento":                      datos.get("estado_agendamiento", "Agendado"),
            "Transcripcion":                              transcripcion,
            "Que genero el agendamiento":                 datos.get("factor_agendamiento", ""),
            "Expectativa de precio del cliente":          datos.get("expectativa_precio", ""),
            "Senales de duda durante la llamada":         datos.get("senales_duda", ""),
            "Brecha precio-expectativa":                  datos.get("brecha_precio_expectativa", ""),
            "Riesgo de no show":                          datos.get("riesgo_no_show", ""),
            "Riesgo comercial identificado":              datos.get("riesgo_comercial", ""),
            "Condiciones especiales mencionadas":         datos.get("condiciones_especiales", "Ninguna"),
            "Recomendacion para el inspector":            datos.get("recomendacion_inspector", ""),
            "Perfil del cliente para el inspector":       datos.get("perfil_inspector", ""),
        }
        print("OK")

        if idx % 10 == 0:
            guardar_progreso(progreso)
            print(f"    [Checkpoint] {idx}/{len(pendientes)}")

    guardar_progreso(progreso)

    # Generar Excel
    print("\n  Generando Excel...")
    wb, ws = crear_excel_vacio()

    MAPA_COLUMNAS = {
        "Nombre de Archivo":                    "Nombre de Archivo",
        "Cliente":                              "Cliente",
        "Contacto Efectivo":                    "Contacto Efectivo",
        "Tiempo de Conversacion":               "Tiempo de Conversación",
        "Actitud Inicial":                      "Actitud Inicial",
        "Vehiculo / Anio":                      "Vehículo / Año",
        "Patente":                              "Patente",
        "Licencia Clase B":                     "Licencia Clase B",
        "Sin Encargo Robo":                     "Sin Encargo Robo",
        "Sin Remate":                           "Sin Remate",
        ">30 Dias Titular":                     ">30 Días Titular",
        "Prenda / Financiera":                  "Prenda / Financiera",
        "Gestion de Objeciones / Dudas":        "Gestión de Objeciones / Dudas",
        "Incentivo / Gancho Comercial":         "Incentivo / Gancho Comercial",
        "Estado / Agendamiento":                "Estado / Agendamiento",
        "Transcripcion":                        "Transcripción",
        "Que genero el agendamiento":           "¿Qué generó el agendamiento?",
        "Expectativa de precio del cliente":    "Expectativa de precio del cliente",
        "Senales de duda durante la llamada":   "Señales de duda durante la llamada",
        "Brecha precio-expectativa":            "Brecha precio-expectativa",
        "Riesgo de no show":                    "Riesgo de no show",
        "Riesgo comercial identificado":        "Riesgo comercial identificado",
        "Condiciones especiales mencionadas":   "Condiciones especiales mencionadas",
        "Recomendacion para el inspector":      "Recomendación para el inspector",
        "Perfil del cliente para el inspector": "Perfil del cliente para el inspector",
    }

    for fila_num, (archivo, datos) in enumerate(progreso.items(), 2):
        row_data = {MAPA_COLUMNAS.get(k, k): v for k, v in datos.items()}
        row_data["Nombre de Archivo"] = archivo
        escribir_fila(ws, fila_num, row_data)

    ajustar_anchos(ws)
    wb.save(EXCEL_SALIDA)

    total = len(progreso)
    print(f"\n{'=' * 60}")
    print(f"  COMPLETADO")
    print(f"  Excel: {EXCEL_SALIDA}")
    print(f"  Total registros: {total}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
