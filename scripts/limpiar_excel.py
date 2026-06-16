"""
Limpia el Excel final:
- Reprocesa las filas con error (transcripcion invalida)
- Elimina filas sin ningun registro util
- Guarda version limpia en datos/
"""
import json, re, os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import anthropic
import whisper

LLAMADAS_DIR = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\llamadas")
DATOS_DIR    = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos")
EXCEL_ORIGEN = DATOS_DIR / "Analisis_Llamadas_Completo.xlsx"
EXCEL_SALIDA = DATOS_DIR / "Analisis_Llamadas_Limpio.xlsx"

COLUMNAS_NUEVAS = [
    "Transcripción",
    "¿Preguntó por precio?",
    "Precio mencionado ($)",
    "Fricción en el precio",
    "Comentario del cliente sobre precio",
    "Motivo de venta",
    "¿Ya cotizó en otro lado?",
    "Señales de decisión de compra",
    "Sugerencias del cliente",
    "Ficha Inspector",
]

PROMPT = """Eres un analista experto en contact center de compra de vehiculos usados.
Esta llamada termino en UNA COMPRA EXITOSA: el cliente agendo una cita de inspeccion.
Extrae los factores que determinaron esa compra desde el primer contacto.

TRANSCRIPCION:
{transcripcion}

Responde UNICAMENTE con JSON valido, sin texto adicional:
{{
    "pregunto_precio": "Si" o "No",
    "precio_mencionado": "monto en pesos o No mencionado",
    "friccion_precio": "Acepta" o "No le gusta" o "N/A",
    "comentario_precio": "cita textual del cliente sobre el precio, o N/A",
    "motivo_venta": "Necesita liquidez / Cambio de auto / Urgencia alta / Sin urgencia / No menciona",
    "cotizo_otro_lado": "Si" o "No" o "No menciona",
    "senales_compra": "factores concretos que determinaron la decision (max 120 palabras)",
    "sugerencias_cliente": "que pide o propone el cliente, o Ninguna",
    "ficha_inspector": "ACTITUD del cliente, URGENCIA de venta, PRECIO esperado, ESTADO del auto, ADVERTENCIAS para la cita. Max 150 palabras."
}}"""


def es_valida(transcripcion):
    if not transcripcion:
        return False
    t = str(transcripcion).strip()
    if t in ("", "Sin transcripción", "Sin transcripcion"):
        return False
    if t.startswith("Error") or t.startswith("[ERROR"):
        return False
    return True


def analizar(client, transcripcion, archivo):
    texto = transcripcion[:9000]
    try:
        response = client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=1200,
            messages=[{"role": "user", "content": PROMPT.format(transcripcion=texto)}]
        )
        raw = response.content[0].text.strip()
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        data = json.loads(raw)
        return {
            COLUMNAS_NUEVAS[0]:  transcripcion,
            COLUMNAS_NUEVAS[1]:  data.get("pregunto_precio", "N/A"),
            COLUMNAS_NUEVAS[2]:  data.get("precio_mencionado", "N/A"),
            COLUMNAS_NUEVAS[3]:  data.get("friccion_precio", "N/A"),
            COLUMNAS_NUEVAS[4]:  data.get("comentario_precio", "N/A"),
            COLUMNAS_NUEVAS[5]:  data.get("motivo_venta", "N/A"),
            COLUMNAS_NUEVAS[6]:  data.get("cotizo_otro_lado", "N/A"),
            COLUMNAS_NUEVAS[7]:  data.get("senales_compra", "N/A"),
            COLUMNAS_NUEVAS[8]:  data.get("sugerencias_cliente", "N/A"),
            COLUMNAS_NUEVAS[9]:  data.get("ficha_inspector", "N/A"),
        }
    except Exception as e:
        print(f"  ERROR Claude para {archivo}: {e}")
        return None


def main():
    print("=" * 55)
    print("  LIMPIEZA Y REGENERACION DEL EXCEL")
    print("=" * 55)

    # Cargar Excel origen
    wb_in = openpyxl.load_workbook(EXCEL_ORIGEN)
    ws_in = wb_in.active
    headers = [cell.value for cell in ws_in[1]]
    idx_trans = next((i for i, h in enumerate(headers) if h and "Transcri" in str(h)), None)
    idx_arch  = 0

    print(f"\n[1] Excel origen: {ws_in.max_row - 1} filas, {len(headers)} columnas")

    # Clasificar filas
    filas_ok = []
    filas_error = []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        trans = row[idx_trans] if idx_trans is not None else None
        if es_valida(trans):
            filas_ok.append(row)
        else:
            filas_error.append(row)

    print(f"    Filas validas:      {len(filas_ok)}")
    print(f"    Filas con error:    {len(filas_error)}")
    for r in filas_error:
        print(f"      -> {r[idx_arch]}")

    # Reprocesar las filas con error
    if filas_error:
        print(f"\n[2] Reprocesando {len(filas_error)} fila(s)...")
        modelo = whisper.load_model("small")
        client = anthropic.Anthropic()

        for row in filas_error:
            archivo = row[idx_arch]
            ruta = LLAMADAS_DIR / archivo
            if not ruta.exists():
                print(f"  SKIP: {archivo} no existe en llamadas/")
                continue

            print(f"  Transcribiendo {archivo}...")
            result = modelo.transcribe(str(ruta), language="es", task="transcribe", fp16=False)
            transcripcion = result["text"].strip()
            print(f"  {len(transcripcion)} chars -> Analizando con Claude...")

            nuevo_analisis = analizar(client, transcripcion, archivo)
            if nuevo_analisis is None:
                print(f"  SKIP: no se pudo analizar {archivo}")
                continue

            # Construir fila actualizada
            nueva_fila = list(row)
            for col_nombre, valor in nuevo_analisis.items():
                if col_nombre in headers:
                    nueva_fila[headers.index(col_nombre)] = valor
            filas_ok.append(tuple(nueva_fila))
            print(f"  OK: {archivo} reprocesado")
    else:
        print("\n[2] Sin filas con error, no se reprocesa nada.")

    # Crear Excel limpio
    print(f"\n[3] Generando Excel limpio con {len(filas_ok)} filas...")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Analisis Completo"
    ws_out.freeze_panes = "A2"

    # Encabezados
    for j, h in enumerate(headers, 1):
        cell = ws_out.cell(row=1, column=j, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        es_nueva = h in COLUMNAS_NUEVAS
        cell.fill = PatternFill("solid", fgColor="2E75B6" if es_nueva else "1F4E79")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    # Datos
    for i, row in enumerate(filas_ok, start=2):
        for j, valor in enumerate(row, 1):
            cell = ws_out.cell(row=i, column=j, value=valor)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Anchos de columna
    for col in ws_out.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value).split("\n")[0]), 60))
        ws_out.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

    wb_out.save(EXCEL_SALIDA)
    print(f"\n{'=' * 55}")
    print(f"  COMPLETADO")
    print(f"  Excel: {EXCEL_SALIDA}")
    print(f"  Total filas: {len(filas_ok)}")
    print(f"{'=' * 55}")


if __name__ == "__main__":
    main()
