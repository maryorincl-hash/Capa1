import openpyxl
from collections import Counter

wb = openpyxl.load_workbook(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos\Analisis_Llamadas_Final.xlsx")
ws = wb.active
headers = [cell.value for cell in ws[1]]
rows = list(ws.iter_rows(min_row=2, values_only=True))

def col(nombre):
    for i, h in enumerate(headers):
        if h and nombre.lower() in str(h).lower():
            return i
    return None

cols_analizar = [
    "Actitud Inicial",
    "Motivo de venta",
    "Pregunt",
    "Fricci",
    "cotiz",
    "Licencia",
    "Encargo",
    "Remate",
    "Titular",
    "Prenda",
]

for nombre in cols_analizar:
    idx = col(nombre)
    if idx is None:
        print(f"{nombre}: NO ENCONTRADA")
        continue
    valores = [str(r[idx]).strip() for r in rows if r[idx]]
    conteo = Counter(valores)
    print(f"\n[{headers[idx]}]  ({len(valores)} registros)")
    for v, c in conteo.most_common(10):
        barra = "#" * c
        print(f"  {v[:45]:<45} {c:3}  {barra}")

# Muestra de sugerencias y senales
print("\n\n--- MUESTRA SUGERENCIAS DEL CLIENTE (primeras 15) ---")
idx_sug = col("Sugerencias")
for r in rows[:50]:
    v = r[idx_sug] if idx_sug is not None else ""
    if v and str(v).strip() not in ("Ninguna", "N/A", ""):
        print(f"  - {str(v)[:120]}")

print("\n--- MUESTRA SENALES DE COMPRA (primeras 10) ---")
idx_sen = col("Senales")
for r in rows[:20]:
    v = r[idx_sen] if idx_sen is not None else ""
    if v:
        print(f"  - {str(v)[:120]}")
