import json, re, os
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment
import anthropic
import whisper

LLAMADAS_DIR = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\llamadas")
DATOS_DIR    = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos")
EXCEL_SALIDA = DATOS_DIR / "Analisis_Llamadas_Completo.xlsx"
ARCHIVO      = "56979221684.mp3"

PROMPT = """Eres un analista experto en contact center de compra de vehiculos usados.
Esta llamada termino en UNA COMPRA EXITOSA: el cliente agendo una cita de inspeccion.
Extrae los factores que determinaron esa compra desde el primer contacto.

TRANSCRIPCION:
{transcripcion}

Responde UNICAMENTE con JSON valido, sin texto adicional:
{{
    "pregunto_precio": "Si" o "No",
    "precio_mencionado": "monto en pesos o No mencionado",
    "friccion_precio": "Acepta" o "No le gusta" o "N/A",
    "comentario_precio": "cita textual del cliente sobre el precio, o N/A",
    "motivo_venta": "Necesita liquidez / Cambio de auto / Urgencia alta / Sin urgencia / No menciona",
    "cotizo_otro_lado": "Si" o "No" o "No menciona",
    "senales_compra": "factores concretos que determinaron la decision (max 120 palabras)",
    "sugerencias_cliente": "que pide o propone el cliente, o Ninguna",
    "ficha_inspector": "ACTITUD del cliente, URGENCIA de venta, PRECIO esperado, ESTADO del auto, ADVERTENCIAS para la cita. Max 150 palabras."
}}"""

print(f"[1] Transcribiendo {ARCHIVO}...")
modelo = whisper.load_model("small")
result = modelo.transcribe(str(LLAMADAS_DIR / ARCHIVO), language="es", task="transcribe", fp16=False)
transcripcion = result["text"].strip()
print(f"    {len(transcripcion)} chars")
print(f"    Muestra: {transcripcion[:200]}\n")

print("[2] Analizando con Claude...")
client = anthropic.Anthropic()
response = client.messages.create(
    model="claude-haiku-4-5",
    max_tokens=1200,
    messages=[{"role": "user", "content": PROMPT.format(transcripcion=transcripcion[:9000])}]
)
raw = response.content[0].text.strip()
raw = re.sub(r"^```json\s*", "", raw)
raw = re.sub(r"\s*```$", "", raw)
data = json.loads(raw)
print("    Resultado:")
for k, v in data.items():
    print(f"      {k}: {str(v)[:100]}")

print("\n[3] Actualizando Excel...")
wb = openpyxl.load_workbook(EXCEL_SALIDA)
ws = wb.active
headers = [cell.value for cell in ws[1]]

fila_target = None
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    if row[0].value == ARCHIVO:
        fila_target = i
        break

if fila_target is None:
    print(f"    WARN: {ARCHIVO} no encontrado en el Excel.")
else:
    campos = {
        "Transcripci":           transcripcion,
        "Pregunt":               data.get("pregunto_precio", "N/A"),
        "Precio mencionado":     data.get("precio_mencionado", "N/A"),
        "Fricci":                data.get("friccion_precio", "N/A"),
        "Comentario":            data.get("comentario_precio", "N/A"),
        "Motivo":                data.get("motivo_venta", "N/A"),
        "cotiz":                 data.get("cotizo_otro_lado", "N/A"),
        "Se":                    data.get("senales_compra", "N/A"),
        "Sugerencias":           data.get("sugerencias_cliente", "N/A"),
        "Ficha":                 data.get("ficha_inspector", "N/A"),
    }
    for j, h in enumerate(headers, 1):
        if not h:
            continue
        for prefijo, valor in campos.items():
            if str(h).startswith(prefijo):
                ws.cell(row=fila_target, column=j, value=valor).alignment = Alignment(wrap_text=True, vertical="top")
                break
    wb.save(EXCEL_SALIDA)
    print(f"    Fila {fila_target} actualizada correctamente en el Excel.")
