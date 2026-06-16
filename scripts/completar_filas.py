"""
Completa las filas que tienen transcripcion pero les faltan las columnas 1-15.
Usa Claude para extraer la informacion desde la transcripcion.
"""
import json, re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import anthropic

LLAMADAS_DIR = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\llamadas")
DATOS_DIR    = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos")
EXCEL_ORIGEN = DATOS_DIR / "Analisis_Llamadas_Limpio.xlsx"
EXCEL_SALIDA = DATOS_DIR / "Analisis_Llamadas_Final.xlsx"

COLUMNAS_ORIGINALES = [
    "Nombre de Archivo",
    "Cliente",
    "Contacto Efectivo",
    "Tiempo de Conversación",
    "Actitud Inicial",
    "Vehículo / Año",
    "Patente",
    "Licencia Clase B",
    "Sin Encargo Robo",
    "Sin Remate",
    ">30 Días Titular",
    "Prenda / Financiera",
    "Gestión de Objeciones / Dudas",
    "Incentivo / Gancho Comercial",
    "Estado / Agendamiento",
]

COLUMNAS_NUEVAS = [
    "Transcripción",
    "¿Preguntó por precio?",
    "Precio mencionado ($)",
    "Fricción en el precio",
    "Comentario del cliente sobre precio",
    "Motivo de venta",
    "¿Ya cotizó en otro lado?",
    "Señales de decisión de compra",
    "Sugerencias del cliente",
    "Ficha Inspector",
]

PROMPT_EXTRACCION = """Eres un analista de contact center de compra de vehiculos usados en Chile.
Esta llamada termino en una COMPRA EXITOSA (el cliente agendo cita de inspeccion).
Extrae la informacion solicitada SOLO desde el contenido de la transcripcion.

TRANSCRIPCION:
{transcripcion}

Responde UNICAMENTE con JSON valido:
{{
    "cliente": "nombre completo del cliente segun la transcripcion, o Desconocido si no se menciona",
    "contacto_efectivo": "Si",
    "tiempo_conversacion": "duracion estimada en formato MM:SS min basandote en la extension del dialogo (ejemplo: 02:15 min)",
    "actitud_inicial": "una o dos palabras: Desconfiado / Apurado / Receptivo / Inseguro / Curioso / Molesto / Tranquilo",
    "vehiculo_anio": "Marca Modelo (AAAA) ejemplo: Toyota Corolla (2018), o No mencionado",
    "patente": "patente en formato chileno si se menciona, o No mencionada",
    "licencia_clase_b": "Si si confirma tener licencia clase B / No si dice no tenerla / No menciona",
    "sin_encargo_robo": "Si si confirma sin encargo de robo / No si hay problemas / No menciona",
    "sin_remate": "Si si confirma sin remate / No si hay problemas / No menciona",
    "dias_titular": "Si si lleva mas de 30 dias como titular / No si menos / No menciona",
    "prenda_financiera": "Si si el auto tiene prenda o credito vigente / No si esta libre / No menciona",
    "gestion_objeciones": "descripcion breve de las dudas u objeciones que tuvo el cliente y como fueron manejadas. Max 100 palabras.",
    "incentivo_gancho": "descripcion del incentivo o gancho comercial que uso el agente para cerrar la cita. Max 80 palabras.",
    "estado_agendamiento": "Agendado: [dia hora (sucursal/ciudad)] segun lo acordado en la llamada. Ejemplo: Agendado: Lunes 10:00 (Huechuraba)"
}}"""


def fila_tiene_datos_originales(row):
    """Retorna True si la fila tiene datos en las primeras 15 columnas."""
    return any(row[i] for i in range(1, 15))


def extraer_con_claude(client, transcripcion, archivo):
    texto = transcripcion[:9000]
    try:
        response = client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=1024,
            messages=[{"role": "user", "content": PROMPT_EXTRACCION.format(transcripcion=texto)}]
        )
        raw = response.content[0].text.strip()
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        data = json.loads(raw)
        return {
            "Cliente":                        data.get("cliente", "Desconocido"),
            "Contacto Efectivo":              data.get("contacto_efectivo", "Sí"),
            "Tiempo de Conversación":         data.get("tiempo_conversacion", "N/A"),
            "Actitud Inicial":                data.get("actitud_inicial", "N/A"),
            "Vehículo / Año":                 data.get("vehiculo_anio", "No mencionado"),
            "Patente":                        data.get("patente", "No mencionada"),
            "Licencia Clase B":               data.get("licencia_clase_b", "No menciona"),
            "Sin Encargo Robo":               data.get("sin_encargo_robo", "No menciona"),
            "Sin Remate":                     data.get("sin_remate", "No menciona"),
            ">30 Días Titular":               data.get("dias_titular", "No menciona"),
            "Prenda / Financiera":            data.get("prenda_financiera", "No menciona"),
            "Gestión de Objeciones / Dudas":  data.get("gestion_objeciones", "N/A"),
            "Incentivo / Gancho Comercial":   data.get("incentivo_gancho", "N/A"),
            "Estado / Agendamiento":          data.get("estado_agendamiento", "Agendado"),
        }
    except Exception as e:
        print(f"  ERROR Claude {archivo}: {e}")
        return None


def main():
    print("=" * 55)
    print("  COMPLETAR FILAS SIN DATOS ORIGINALES")
    print("=" * 55)

    # Cargar Excel
    wb = openpyxl.load_workbook(EXCEL_ORIGEN)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    idx_trans = next((i for i, h in enumerate(headers) if h and "Transcri" in str(h)), None)

    # Identificar filas vacias y filas con datos
    archivos_en_excel = set()
    filas_vacias = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0]:
            archivos_en_excel.add(row[0])
        else:
            filas_vacias.append(i)

    print(f"\n[1] Filas con datos completos: {len(archivos_en_excel)}")
    print(f"    Filas sin datos originales: {len(filas_vacias)}")

    # Identificar los MP3s que no estan en el Excel (sin filename asignado)
    todos_mp3s = sorted(LLAMADAS_DIR.glob("*.mp3"))
    mp3s_nuevos = [p for p in todos_mp3s if p.name not in archivos_en_excel]
    print(f"    MP3s no asignados: {len(mp3s_nuevos)}")

    if len(filas_vacias) != len(mp3s_nuevos):
        print(f"  AVISO: {len(filas_vacias)} filas vacias pero {len(mp3s_nuevos)} MP3s nuevos — se asignan en orden.")

    # Asignar filenames a filas vacias (en orden)
    asignaciones = {}
    for fila, mp3 in zip(filas_vacias, mp3s_nuevos):
        asignaciones[fila] = mp3.name

    print(f"\n[2] Extrayendo informacion con Claude para {len(filas_vacias)} filas...")
    client = anthropic.Anthropic()
    errores = 0

    for idx, (fila_num, archivo) in enumerate(asignaciones.items(), 1):
        row_data = list(ws.iter_rows(min_row=fila_num, max_row=fila_num, values_only=True))[0]
        transcripcion = row_data[idx_trans] if idx_trans is not None else ""

        if not transcripcion or len(str(transcripcion).strip()) < 50:
            print(f"  [{idx}/{len(asignaciones)}] {archivo} -> SKIP (sin transcripcion util)")
            errores += 1
            continue

        print(f"  [{idx}/{len(asignaciones)}] {archivo}...", end=" ", flush=True)
        datos = extraer_con_claude(client, str(transcripcion), archivo)

        if datos is None:
            errores += 1
            continue

        # Escribir en el Excel
        ws.cell(row=fila_num, column=1, value=archivo).alignment = Alignment(wrap_text=True, vertical="top")
        for col_nombre, valor in datos.items():
            if col_nombre in headers:
                j = headers.index(col_nombre) + 1
                ws.cell(row=fila_num, column=j, value=valor).alignment = Alignment(wrap_text=True, vertical="top")

        print("OK")

        # Checkpoint cada 10
        if idx % 10 == 0:
            wb.save(EXCEL_SALIDA)
            print(f"    [Checkpoint] {idx}/{len(asignaciones)}")

    # Guardar final
    # Ajustar anchos
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value).split("\n")[0]), 60))
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

    wb.save(EXCEL_SALIDA)

    print(f"\n{'=' * 55}")
    print(f"  COMPLETADO")
    print(f"  Excel: {EXCEL_SALIDA}")
    print(f"  Filas completadas: {len(filas_vacias) - errores}")
    if errores:
        print(f"  Errores: {errores}")
    print(f"{'=' * 55}")


if __name__ == "__main__":
    main()
