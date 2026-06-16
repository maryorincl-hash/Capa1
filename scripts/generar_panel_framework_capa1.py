# Panel Framework Capa 1 — Llamadas sin agendamiento
# Lee Analisis_Capa1_Framework.xlsx y genera Panel_Framework_Capa1.html
#
# Antes de correr:
#   $env:ANTHROPIC_API_KEY = (Get-ItemProperty "HKCU:\Environment" -Name "ANTHROPIC_API_KEY").ANTHROPIC_API_KEY
#   cd C:\Users\maryorin.vivas\Proyectos\LeadsCompra
#   python scripts\generar_panel_framework_capa1.py

import json, re
from pathlib import Path
from collections import Counter, defaultdict
import openpyxl
import anthropic

DATOS_DIR  = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos")
EXCEL      = DATOS_DIR / "Analisis_Capa1_Framework.xlsx"
HTML_SAL   = DATOS_DIR / "Panel_Framework_Capa1.html"

PALETTE = ["#dc2626","#d97706","#2563eb","#16a34a","#7c3aed","#0891b2","#db2777","#65a30d"]

# ── carga ─────────────────────────────────────────────────────────────────────

def cargar():
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if any(r):
            rows.append(dict(zip(headers, r)))
    return headers, rows

def v(row, key):
    val = row.get(key, "")
    return str(val).strip() if val else ""

# ── normalizar valores con variantes ─────────────────────────────────────────

def norm_tipo_fracaso(s):
    s = s.strip()
    if "Falla" in s and ("Fricci" in s or "Brecha" in s or "Timing" in s):
        return "Falla del Agente + Otro"
    if "Falla" in s:
        return "Falla del Agente"
    if "Fricci" in s:
        return "Fricción Logística"
    if "Brecha" in s:
        return "Brecha de Precio"
    if "Timing" in s:
        return "Timing — Solo Evaluando"
    if "Propiedad" in s or "Decisi" in s:
        return "Falta de Propiedad/Decisión"
    return s

def norm_perfil(s):
    if "Pragm" in s and "Timing" in s: return "Pragmático/Timing"
    if "Pragm" in s and "Frust" in s:  return "Pragmático/Frustrado-Precio"
    if "Pragm" in s:                   return "Pragmático"
    if "Timing" in s:                  return "Timing"
    if "Frust" in s:                   return "Frustrado-Precio"
    if "Descon" in s:                  return "Desconfiado"
    if "Anal" in s:                    return "Analítico"
    if "Emoc" in s:                    return "Emocional"
    return "Otro"

def norm_tono(s):
    if "Rob" in s or "Mon" in s: return "Robótico/Monótono"
    if "Neutro" in s:             return "Neutro"
    if "Experto" in s or "Asesor" in s: return "Experto/Asesor"
    if "Agres" in s or "Presi" in s:   return "Agresivo/Presión"
    return s

def norm_cal(s):
    for x in ["Excelente","Buena","Regular","Deficiente"]:
        if x in s: return x
    return s

def norm_a1(s):
    if "Sí" in s or "Si " in s or s.startswith("Sí") or "conect" in s.lower():
        return "Sí conectó"
    if "Genérico" in s or "Generic" in s or "no mencion" in s.lower():
        return "Genérico"
    if "No se" in s or "no ident" in s.lower():
        return "No se identificó"
    return "Genérico"

def norm_b1(s):
    if "No pregunt" in s or "No hizo" in s or "No realiz" in s:
        return "No preguntó nada"
    if "Parcial" in s: return "Parcialmente"
    if "Sí" in s: return "Sí hizo preguntas"
    return "No preguntó nada"

def norm_cita(s):
    if "Beneficio" in s: return "Beneficio"
    if "Trámite" in s or "Tramite" in s: return "Trámite"
    return "No intentó"

def norm_negativa(s):
    if "Rebatió" in s or "Rebatio" in s: return "Rebatió con argumento"
    if "Aceptó" in s or "Acepto" in s or "rindió" in s: return "Aceptó el no"
    return "No hubo negativa explícita"

def norm_2cierre(s):
    if "Sí" in s[:5]: return "Sí intentó"
    if "No aplicaba" in s: return "No aplicaba"
    return "No intentó"

# ── estadísticas ──────────────────────────────────────────────────────────────

def dist(rows, key, norm_fn=None, top=10):
    vals = [v(r, key) for r in rows if v(r, key)]
    if norm_fn:
        vals = [norm_fn(x) for x in vals]
    return Counter(vals).most_common(top)

def pct(n, total): return round(n*100/total) if total else 0

# ── JS chart helpers ──────────────────────────────────────────────────────────

def jl(items):
    L = json.dumps([i[0] for i in items])
    D = json.dumps([i[1] for i in items])
    return L, D

def bar_h(cid, title, items, color="#2563eb"):
    L, D = jl(items)
    return f"""new Chart(document.getElementById('{cid}'),{{type:'bar',
  data:{{labels:{L},datasets:[{{data:{D},backgroundColor:'{color}cc',borderColor:'{color}',borderWidth:1}}]}},
  options:{{indexAxis:'y',responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{display:false}},title:{{display:true,text:'{title}',font:{{size:12,weight:'bold'}}}}}},
    scales:{{x:{{beginAtZero:true,ticks:{{stepSize:1}}}}}}
  }}}});"""

def donut(cid, title, items, colors=None):
    L, D = jl(items)
    BG = json.dumps((colors or PALETTE)[:len(items)])
    return f"""new Chart(document.getElementById('{cid}'),{{type:'doughnut',
  data:{{labels:{L},datasets:[{{data:{D},backgroundColor:{BG},borderWidth:2}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    plugins:{{legend:{{position:'right',labels:{{font:{{size:10}},boxWidth:12}}}},
      title:{{display:true,text:'{title}',font:{{size:12,weight:'bold'}}}}}}
  }}}});"""

def radar_pilares(cid, cals_a, cals_b, cals_c, total):
    def score(counter):
        pts = {"Excelente":4,"Buena":3,"Regular":2,"Deficiente":1}
        total_n = sum(counter.values())
        if not total_n: return 0
        s = sum(pts.get(k,0)*n for k,n in counter.items())
        return round(s * 25 / total_n)
    sA = score(Counter(dict(cals_a)))
    sB = score(Counter(dict(cals_b)))
    sC = score(Counter(dict(cals_c)))
    return f"""new Chart(document.getElementById('{cid}'),{{type:'radar',
  data:{{labels:['Apertura\\n(Pilar A)','Escucha Activa\\n(Pilar B)','Cierre\\n(Pilar C)'],
    datasets:[{{label:'Score promedio /100',
      data:[{sA},{sB},{sC}],
      backgroundColor:'rgba(220,38,38,0.15)',borderColor:'#dc2626',borderWidth:2,
      pointBackgroundColor:'#dc2626'}}]}},
  options:{{responsive:true,maintainAspectRatio:false,
    scales:{{r:{{min:0,max:100,ticks:{{stepSize:25,font:{{size:9}}}},
      pointLabels:{{font:{{size:11,weight:'bold'}}}}}}}},
    plugins:{{legend:{{position:'bottom',labels:{{font:{{size:10}}}}}}}}
  }}}});"""

# ── síntesis Claude ───────────────────────────────────────────────────────────

def sintetizar(client, resumen_datos):
    prompt = f"""Eres analista senior de contact center de compra de autos en Chile.
Analiza estos datos de 97 llamadas donde el cliente atendió pero NO agendó cita:

{resumen_datos}

Escribe en HTML (sin <html>/<body>/<head>) un análisis ejecutivo con:
1. El diagnóstico principal en 2-3 oraciones (qué está fallando estructuralmente)
2. Los 3 hallazgos más críticos en bullets <li>
3. Las 3 acciones prioritarias de coaching en bullets <li>
4. Una frase de cierre sobre el potencial de mejora

Usa etiquetas <h4>, <ul>, <li>, <p>, <strong>. Máximo 350 palabras. Sin markdown."""
    try:
        r = client.messages.create(model="claude-haiku-4-5", max_tokens=900,
            messages=[{"role":"user","content":prompt}])
        return r.content[0].text.strip()
    except Exception as e:
        return f"<p style='color:#ef4444'>Error síntesis: {e}</p>"

# ── fichas JS ─────────────────────────────────────────────────────────────────

def color_cal(cal):
    return {"Excelente":"#16a34a","Buena":"#2563eb","Regular":"#d97706","Deficiente":"#dc2626"}.get(cal,"#64748b")

def color_tipo(t):
    m = {"Falla del Agente":"#dc2626","Falla del Agente + Otro":"#f97316",
         "Fricción Logística":"#d97706","Brecha de Precio":"#7c3aed",
         "Timing — Solo Evaluando":"#0891b2","Falta de Propiedad/Decisión":"#64748b"}
    return m.get(t,"#64748b")

def fichas_js(rows):
    fichas = []
    for r in rows:
        tip = norm_tipo_fracaso(v(r,"Tipificación de Fracaso")) if v(r,"Tipificación de Fracaso") else "Sin datos"
        cal_a = norm_cal(v(r,"A · Calificación Apertura"))
        cal_b = norm_cal(v(r,"B · Calificación Escucha"))
        cal_c = norm_cal(v(r,"C · Calificación Cierre"))
        perfil = norm_perfil(v(r,"Actitud / Perfil del Cliente")) if v(r,"Actitud / Perfil del Cliente") else "—"
        fichas.append({
            "id": v(r,"Nombre de Archivo").replace(".mp3",""),
            "cliente": v(r,"Cliente") or "Desconocido",
            "vehiculo": v(r,"Vehículo / Año") or "No mencionado",
            "tiempo": v(r,"Tiempo de Conversación") or "—",
            "tipo": tip,
            "colorTipo": color_tipo(tip),
            "perfil": perfil,
            "calA": cal_a, "colorA": color_cal(cal_a),
            "calB": cal_b, "colorB": color_cal(cal_b),
            "calC": cal_c, "colorC": color_cal(cal_c),
            "objecion": (v(r,"Objeción Principal (textual)") or "")[:160],
            "coaching": (v(r,"Recomendación de Coaching") or "")[:200],
            "detalle": (v(r,"Detalle del Fracaso") or "")[:200],
            "tono": norm_tono(v(r,"A3 · Tono Inicial del Agente")) if v(r,"A3 · Tono Inicial del Agente") else "—",
            "cita": norm_cita(v(r,"C1 · Cita Vendida Como")) if v(r,"C1 · Cita Vendida Como") else "—",
        })
    return json.dumps(fichas, ensure_ascii=False)

# ── build HTML ────────────────────────────────────────────────────────────────

def build(headers, rows, sintesis):
    N = len(rows)

    # distribuciones normalizadas
    d_tipo   = dist(rows, "Tipificación de Fracaso",   norm_tipo_fracaso)
    d_perfil = dist(rows, "Actitud / Perfil del Cliente", norm_perfil)
    d_tono   = dist(rows, "A3 · Tono Inicial del Agente", norm_tono)
    d_calA   = dist(rows, "A · Calificación Apertura",  norm_cal)
    d_calB   = dist(rows, "B · Calificación Escucha",   norm_cal)
    d_calC   = dist(rows, "C · Calificación Cierre",    norm_cal)
    d_a1     = dist(rows, "A1 · Velocidad de Contexto", norm_a1)
    d_b1     = dist(rows, "B1 · Preguntas de Descubrimiento", norm_b1)
    d_cita   = dist(rows, "C1 · Cita Vendida Como",     norm_cita)
    d_neg    = dist(rows, "C2 · Manejo Primera Negativa", norm_negativa)
    d_2c     = dist(rows, "C3 · Segundo Cierre Intentado", norm_2cierre)
    d_precio = dist(rows, "Tipo de Precio")

    # clústeres consolidados
    tipo_map = defaultdict(int)
    for r in rows:
        t = norm_tipo_fracaso(v(r,"Tipificación de Fracaso")) if v(r,"Tipificación de Fracaso") else "Sin datos"
        tipo_map[t] += 1

    falla_pura   = tipo_map.get("Falla del Agente",0)
    falla_mix    = tipo_map.get("Falla del Agente + Otro",0)
    friccion     = tipo_map.get("Fricción Logística",0)
    brecha       = tipo_map.get("Brecha de Precio",0)
    timing       = tipo_map.get("Timing — Solo Evaluando",0)
    sin_decision = tipo_map.get("Falta de Propiedad/Decisión",0)
    total_falla  = falla_pura + falla_mix

    # charts JS
    charts = "\n".join([
        bar_h("chTipo","Tipificación de Fracaso", d_tipo, "#dc2626"),
        donut("chPerfil","Perfil del Cliente", d_perfil),
        donut("chCalA","Calificación Pilar A", d_calA, ["#16a34a","#2563eb","#d97706","#dc2626"]),
        donut("chCalB","Calificación Pilar B", d_calB, ["#16a34a","#2563eb","#d97706","#dc2626"]),
        donut("chCalC","Calificación Pilar C", d_calC, ["#16a34a","#2563eb","#d97706","#dc2626"]),
        bar_h("chTono","Tono del Agente", d_tono, "#7c3aed"),
        bar_h("chA1","Velocidad de Contexto (A1)", d_a1, "#15803d"),
        bar_h("chB1","Preguntas de Descubrimiento (B1)", d_b1, "#1d4ed8"),
        bar_h("chCita","Cita Vendida Como (C1)", d_cita, "#7c3aed"),
        bar_h("chNeg","Manejo Primera Negativa (C2)", d_neg, "#7c3aed"),
        bar_h("ch2c","Segundo Cierre (C3)", d_2c, "#7c3aed"),
        radar_pilares("chRadar", d_calA, d_calB, d_calC, N),
    ])

    fichas_data = fichas_js(rows)

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Framework Capa 1 — Llamadas sin Agendamiento</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#f1f5f9;color:#1e293b}}
header{{background:linear-gradient(135deg,#7f1d1d,#dc2626);color:#fff;padding:22px 32px}}
header h1{{font-size:1.3rem;font-weight:800}}
header p{{font-size:.82rem;opacity:.85;margin-top:5px}}
.banner{{background:#fef2f2;border-left:5px solid #dc2626;padding:10px 24px;
  font-size:.83rem;color:#7f1d1d;display:flex;gap:16px;flex-wrap:wrap;align-items:center}}
.tabs{{display:flex;gap:4px;background:#fff;padding:10px 20px;border-bottom:1px solid #e2e8f0;flex-wrap:wrap}}
.tab{{padding:8px 16px;border-radius:6px;cursor:pointer;font-size:.83rem;font-weight:600;
  color:#64748b;border:1px solid transparent}}
.tab:hover{{background:#f1f5f9}}.tab.active{{background:#dc2626;color:#fff}}
.sec{{display:none;padding:22px;max-width:1440px;margin:0 auto}}.sec.active{{display:block}}
.kpi-row{{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px;margin-bottom:18px}}
.kpi{{background:#fff;border-radius:10px;padding:14px;text-align:center;box-shadow:0 1px 4px #0001;border-top:4px solid transparent}}
.kpi .num{{font-size:1.9rem;font-weight:800}}.kpi .lbl{{font-size:.71rem;color:#64748b;margin-top:2px;line-height:1.3}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px}}
.g3{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:14px}}
.g4{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:14px}}
.card{{background:#fff;border-radius:10px;padding:18px;box-shadow:0 1px 4px #0001}}
.card h3{{font-size:.88rem;font-weight:700;margin-bottom:12px;color:#1e293b;
  border-bottom:2px solid #dc2626;padding-bottom:5px}}
.card canvas{{max-height:220px}}
.insight{{border-left:5px solid;border-radius:0 10px 10px 0;padding:14px 18px;margin-bottom:12px}}
.insight h3{{font-size:.88rem;font-weight:700;margin-bottom:6px}}
.insight p,.insight li{{font-size:.83rem;line-height:1.65}}
.insight ul{{padding-left:16px}}
.cluster-grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:12px;margin-bottom:16px}}
.cluster{{background:#fff;border-radius:10px;box-shadow:0 1px 4px #0001;overflow:hidden}}
.cluster-head{{padding:14px 16px;color:#fff}}
.cluster-head .pct{{font-size:2rem;font-weight:900;opacity:.9}}
.cluster-head .name{{font-size:.88rem;font-weight:700;margin-top:2px}}
.cluster-body{{padding:12px 16px;font-size:.8rem;line-height:1.6;color:#374151}}
.cluster-body strong{{color:#1e293b}}
.pilar-row{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:14px}}
.pilar{{background:#fff;border-radius:10px;padding:16px;box-shadow:0 1px 4px #0001}}
.pilar-tag{{display:inline-block;padding:3px 10px;border-radius:99px;font-size:.72rem;
  font-weight:700;color:#fff;margin-bottom:10px}}
.pilar h3{{font-size:.88rem;font-weight:700;margin-bottom:10px}}
.pilar canvas{{max-height:180px}}
.bar-stat{{display:flex;align-items:center;gap:8px;margin-bottom:6px;font-size:.8rem}}
.bar-fill{{height:10px;border-radius:5px;min-width:4px}}
.synthesis{{background:#fff;border-radius:10px;padding:20px;box-shadow:0 1px 4px #0001;
  line-height:1.65;font-size:.85rem}}
.synthesis h4{{color:#dc2626;margin:12px 0 5px;font-size:.88rem}}
.synthesis ul{{padding-left:16px}}.synthesis li{{margin-bottom:4px}}
.fichas-bar{{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap;align-items:center}}
.fichas-bar input,.fichas-bar select{{padding:8px 12px;border:1px solid #cbd5e1;
  border-radius:6px;font-size:.82rem}}
.fichas-bar input{{flex:1;min-width:180px}}
.fgrid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:12px}}
.fc{{background:#fff;border-radius:10px;box-shadow:0 1px 4px #0001;overflow:hidden;font-size:.8rem}}
.fc-head{{padding:10px 13px;background:#f8fafc;border-bottom:1px solid #e2e8f0;display:flex;
  justify-content:space-between;align-items:flex-start;gap:6px}}
.fc-name{{font-weight:700;font-size:.86rem}}.fc-sub{{color:#64748b;font-size:.75rem}}
.badge{{padding:2px 8px;border-radius:99px;font-size:.7rem;font-weight:700;color:#fff;white-space:nowrap}}
.fc-body{{padding:10px 13px;color:#374151;line-height:1.5}}.fc-body p{{margin-bottom:5px}}
.fc-foot{{padding:8px 13px;background:#f8fafc;border-top:1px solid #e2e8f0;
  display:flex;gap:6px;flex-wrap:wrap}}
.cal-badge{{padding:2px 7px;border-radius:4px;font-size:.69rem;font-weight:700;color:#fff}}
.hidden{{display:none!important}}
@media(max-width:900px){{.g2,.g3,.g4,.pilar-row{{grid-template-columns:1fr}}}}
</style>
</head>
<body>

<header>
  <h1>Framework Capa 1 — Llamadas sin Agendamiento · Clicar.cl</h1>
  <p>Análisis con Framework Comercial de 3 Pilares · {N} llamadas · ABR-JUN 2025 · Cliente atendió pero NO agendó cita</p>
</header>

<div class="banner">
  📊 <strong>{N} llamadas analizadas</strong>
  &nbsp;·&nbsp; 🔴 <strong>{total_falla} ({pct(total_falla,N)}%)</strong> Falla del Agente
  &nbsp;·&nbsp; 🕐 <strong>{timing} ({pct(timing,N)}%)</strong> Timing/Solo Evaluando
  &nbsp;·&nbsp; 🚗 <strong>{friccion} ({pct(friccion,N)}%)</strong> Fricción Logística
  &nbsp;·&nbsp; 💰 <strong>{brecha} ({pct(brecha,N)}%)</strong> Brecha de Precio
  &nbsp;·&nbsp; 📋 <strong>{sin_decision} ({pct(sin_decision,N)}%)</strong> Sin Decisión
</div>

<div class="tabs">
  <div class="tab active" onclick="show('s0',this)">🔍 Resumen Ejecutivo</div>
  <div class="tab" onclick="show('s1',this)">🎯 Clústeres</div>
  <div class="tab" onclick="show('s2',this)">📊 Pilares A · B · C</div>
  <div class="tab" onclick="show('s3',this)">🚨 Tipificaciones</div>
  <div class="tab" onclick="show('s4',this)">👤 Perfiles</div>
  <div class="tab" onclick="show('s5',this)">📋 Fichas</div>
</div>

<!-- ─── TAB 0 RESUMEN ──────────────────────────────────────────────────── -->
<div id="s0" class="sec active">
  <div class="kpi-row">
    <div class="kpi" style="border-color:#dc2626"><div class="num" style="color:#dc2626">{N}</div><div class="lbl">Llamadas sin cita</div></div>
    <div class="kpi" style="border-color:#dc2626"><div class="num" style="color:#dc2626">{pct(total_falla,N)}%</div><div class="lbl">Falla del Agente</div></div>
    <div class="kpi" style="border-color:#d97706"><div class="num" style="color:#d97706">{pct(next((n for v,n in d_tono if "Rob" in v),0),N)}%</div><div class="lbl">Tono Robótico</div></div>
    <div class="kpi" style="border-color:#7c3aed"><div class="num" style="color:#7c3aed">{pct(next((n for v,n in d_b1 if "No" in v),0),N)}%</div><div class="lbl">Sin preguntas descubrimiento</div></div>
    <div class="kpi" style="border-color:#7c3aed"><div class="num" style="color:#7c3aed">{pct(next((n for v,n in d_2c if "No intentó" in v),0),N)}%</div><div class="lbl">Sin segundo cierre</div></div>
    <div class="kpi" style="border-color:#7c3aed"><div class="num" style="color:#7c3aed">{pct(next((n for v,n in d_cita if "No intentó" in v),0),N)}%</div><div class="lbl">No vendió la cita</div></div>
    <div class="kpi" style="border-color:#dc2626"><div class="num" style="color:#dc2626">{pct(next((n for v,n in d_calC if "Deficiente" in v),0),N)}%</div><div class="lbl">Cierre Deficiente</div></div>
    <div class="kpi" style="border-color:#dc2626"><div class="num" style="color:#dc2626">{pct(next((n for v,n in d_calB if "Deficiente" in v),0),N)}%</div><div class="lbl">Escucha Deficiente</div></div>
  </div>

  <div class="g2">
    <div class="card" style="height:300px"><h3>Radar de Desempeño — Pilares A · B · C</h3><canvas id="chRadar"></canvas></div>
    <div class="card"><h3>Síntesis Ejecutiva</h3>
      <div class="synthesis">{sintesis}</div>
    </div>
  </div>

  <div class="g2">
    <div class="card" style="height:280px"><h3>Tipificación de Fracaso</h3><canvas id="chTipo"></canvas></div>
    <div class="card" style="height:280px"><h3>Perfil del Cliente</h3><canvas id="chPerfil"></canvas></div>
  </div>
</div>

<!-- ─── TAB 1 CLÚSTERES ───────────────────────────────────────────────── -->
<div id="s1" class="sec">
  <div class="insight" style="border-color:#dc2626;background:#fef2f2;margin-bottom:20px">
    <h3 style="color:#991b1b">Diagnóstico: 3 de cada 4 pérdidas son prevenibles con coaching</h3>
    <p style="color:#7f1d1d">El {pct(total_falla,N)}% de las llamadas fallidas se explica por deficiencias en la ejecución del agente, no por el perfil del cliente. El tono robótico, la ausencia de preguntas de descubrimiento y la falta de un segundo cierre son los patrones dominantes.</p>
  </div>

  <div class="cluster-grid">
    <div class="cluster">
      <div class="cluster-head" style="background:#dc2626">
        <div class="pct">{pct(total_falla,N)}%</div>
        <div class="name">CLUSTER 1 — Falla del Agente</div>
      </div>
      <div class="cluster-body">
        <strong>{total_falla} llamadas</strong><br><br>
        El cliente estaba disponible y en muchos casos abierto, pero el agente no ejecutó correctamente:<br><br>
        • <strong>{pct(next((n for v,n in d_tono if "Rob" in v),0),N)}%</strong> tono robótico/monótono<br>
        • <strong>{pct(next((n for v,n in d_b1 if "No" in v),0),N)}%</strong> sin preguntas de descubrimiento<br>
        • <strong>{pct(next((n for v,n in d_2c if "No intentó" in v),0),N)}%</strong> sin segundo cierre<br>
        • <strong>{pct(next((n for v,n in d_cita if "No intentó" in v),0),N)}%</strong> no intentó vender la cita<br>
        • <strong>{pct(next((n for v,n in d_neg if "Aceptó" in v),0),N)}%</strong> aceptó el primer "no" sin rebatir<br><br>
        <em>Acción: Coaching en apertura personalizada, técnicas de cierre y manejo de objeciones.</em>
      </div>
    </div>

    <div class="cluster">
      <div class="cluster-head" style="background:#0891b2">
        <div class="pct">{pct(timing,N)}%</div>
        <div class="name">CLUSTER 2 — Timing / Solo Evaluando</div>
      </div>
      <div class="cluster-body">
        <strong>{timing} llamadas</strong><br><br>
        El cliente tasó su auto para conocer el valor de mercado, sin urgencia real de vender. No es pérdida definitiva — es lead en maduración.<br><br>
        • No tiene planes inmediatos de venta<br>
        • Estaba "indexando su patrimonio"<br>
        • El agente no detectó esto por falta de preguntas de descubrimiento<br><br>
        <em>Acción: Detectar temprano con pregunta "¿Para cuándo estás pensando vender?" y derivar a seguimiento programado.</em>
      </div>
    </div>

    <div class="cluster">
      <div class="cluster-head" style="background:#d97706">
        <div class="pct">{pct(friccion,N)}%</div>
        <div class="name">CLUSTER 3 — Fricción Logística</div>
      </div>
      <div class="cluster-body">
        <strong>{friccion} llamadas</strong><br><br>
        El cliente tenía interés en vender pero barreras prácticas: sucursal lejos, horarios incompatibles, no puede mover el auto.<br><br>
        • Hay intención real de venta<br>
        • El obstáculo es logístico, no de precio ni de interés<br>
        • En muchos casos el agente no ofreció alternativas de sucursal u horario<br><br>
        <em>Acción: Entrenar en C4 — ofrecer siempre múltiples sucursales y horarios antes de dar por perdida la cita.</em>
      </div>
    </div>

    <div class="cluster">
      <div class="cluster-head" style="background:#7c3aed">
        <div class="pct">{pct(brecha,N)}%</div>
        <div class="name">CLUSTER 4 — Brecha de Precio</div>
      </div>
      <div class="cluster-body">
        <strong>{brecha} llamadas</strong><br><br>
        El cliente esperaba recibir más dinero del que ofrece la plataforma. El precio web quedó muy por debajo de su expectativa.<br><br>
        • El cliente verbalizó que el precio no alcanza<br>
        • En varios casos el agente no argumentó el valor diferencial de la inspección presencial<br>
        • La cita se presentó como trámite en lugar de como oportunidad de ajuste de precio<br><br>
        <em>Acción: Argumentar que la inspección puede mejorar la oferta según estado real del auto.</em>
      </div>
    </div>

    <div class="cluster">
      <div class="cluster-head" style="background:#64748b">
        <div class="pct">{pct(sin_decision,N)}%</div>
        <div class="name">CLUSTER 5 — Sin Decisión</div>
      </div>
      <div class="cluster-body">
        <strong>{sin_decision} llamadas</strong><br><br>
        El cliente no es el tomador de decisión o hay una traba objetiva: auto de familiar, prenda/financiera activa, o co-propietario.<br><br>
        • Obstáculo estructural, difícil de superar en la misma llamada<br>
        • Parte de estos casos podrían recuperarse con seguimiento al titular real<br><br>
        <em>Acción: Solicitar datos del titular real y programar re-contacto directo.</em>
      </div>
    </div>
  </div>
</div>

<!-- ─── TAB 2 PILARES ─────────────────────────────────────────────────── -->
<div id="s2" class="sec">
  <div class="kpi-row">
    <div class="kpi" style="border-color:#15803d">
      <div class="num" style="color:#15803d">{pct(next((n for v,n in d_calA if "Deficiente" in v),0)+next((n for v,n in d_calA if "Regular" in v),0),N)}%</div>
      <div class="lbl">Apertura: Def+Reg</div>
    </div>
    <div class="kpi" style="border-color:#1d4ed8">
      <div class="num" style="color:#1d4ed8">{pct(next((n for v,n in d_calB if "Deficiente" in v),0),N)}%</div>
      <div class="lbl">Escucha: Deficiente</div>
    </div>
    <div class="kpi" style="border-color:#7c3aed">
      <div class="num" style="color:#7c3aed">{pct(next((n for v,n in d_calC if "Deficiente" in v),0),N)}%</div>
      <div class="lbl">Cierre: Deficiente</div>
    </div>
    <div class="kpi" style="border-color:#dc2626">
      <div class="num" style="color:#dc2626">{pct(next((n for v,n in d_tono if "Rob" in v),0)+next((n for v,n in d_tono if "Neutro" in v and "Rob" in v),0),N)}%</div>
      <div class="lbl">Tono Robótico</div>
    </div>
  </div>

  <div class="pilar-row">
    <div class="pilar">
      <span class="pilar-tag" style="background:#15803d">PILAR A — Apertura</span>
      <div style="height:160px;margin-bottom:14px"><canvas id="chCalA"></canvas></div>
      <div style="height:150px"><canvas id="chA1"></canvas></div>
      <div style="height:150px;margin-top:14px"><canvas id="chTono"></canvas></div>
    </div>
    <div class="pilar">
      <span class="pilar-tag" style="background:#1d4ed8">PILAR B — Escucha Activa</span>
      <div style="height:160px;margin-bottom:14px"><canvas id="chCalB"></canvas></div>
      <div style="height:150px"><canvas id="chB1"></canvas></div>
      <div class="insight" style="border-color:#1d4ed8;background:#eff6ff;margin-top:14px">
        <h3 style="color:#1e40af;font-size:.82rem">Hallazgo crítico</h3>
        <p style="color:#1e3a5f;font-size:.8rem">El <strong>{pct(next((n for v,n in d_calB if "Deficiente" in v),0),N)}%</strong> de los agentes tuvo escucha activa deficiente.
        Sin preguntas de descubrimiento, el agente no sabe si el cliente quiere renovar, necesita dinero o solo estaba curioseando — y aplica el mismo guión para todos.</p>
      </div>
    </div>
    <div class="pilar">
      <span class="pilar-tag" style="background:#7c3aed">PILAR C — Cierre</span>
      <div style="height:160px;margin-bottom:14px"><canvas id="chCalC"></canvas></div>
      <div style="height:150px"><canvas id="chCita"></canvas></div>
      <div style="height:150px;margin-top:14px"><canvas id="chNeg"></canvas></div>
      <div style="height:130px;margin-top:14px"><canvas id="ch2c"></canvas></div>
    </div>
  </div>
</div>

<!-- ─── TAB 3 TIPIFICACIONES ──────────────────────────────────────────── -->
<div id="s3" class="sec">
  <div class="g2">
    <div class="card" style="height:320px"><h3>Distribución de Fracasos</h3><canvas id="chTipo2"></canvas></div>
    <div>
"""
    for tipo, n in d_tipo:
        col = color_tipo(tipo)
        pct_v = pct(n, N)
        html += f"""      <div class="insight" style="border-color:{col};background:{col}12;margin-bottom:10px">
        <h3 style="color:{col}">{tipo} — {n} llamadas ({pct_v}%)</h3>
        <div class="bar-stat"><div class="bar-fill" style="width:{pct_v*3}px;background:{col}"></div><span style="font-size:.78rem;color:#64748b">{pct_v}% del total</span></div>
      </div>\n"""
    html += f"""    </div>
  </div>
</div>

<!-- ─── TAB 4 PERFILES ────────────────────────────────────────────────── -->
<div id="s4" class="sec">
  <div class="g2">
    <div class="card" style="height:320px"><h3>Perfil de Comunicación del Cliente</h3><canvas id="chPerfil2"></canvas></div>
    <div>
"""
    perfil_desc = {
        "Pragmático": "Directo, quiere datos concretos y poco tiempo. Responde a argumentos de valor claro. Con este perfil, el agente debe ir al grano: precio garantizado, 20 minutos, sin compromiso.",
        "Timing": "Solo estaba indexando el valor de su auto. No hay urgencia real de venta. El agente debe detectarlo con una pregunta de descubrimiento y no forzar el cierre.",
        "Frustrado-Precio": "El precio web quedó muy por debajo de su expectativa. Necesita que el agente explique que la inspección presencial puede ajustar la oferta hacia arriba.",
        "Pragmático/Timing": "Perfil mixto: directo pero sin urgencia. El agente debe preguntar el timeline de venta y programar un re-contacto si no hay urgencia inmediata.",
        "Pragmático/Frustrado-Precio": "Directo y con expectativa de precio alta. Requiere argumento concreto sobre cómo la inspección puede mejorar la oferta.",
        "Desconfiado": "Escéptico y resistente. Necesita validación y confianza antes de comprometerse. El tono robótico agrava su resistencia.",
        "Analítico": "Hace preguntas, quiere entender el proceso completo. El agente debe ser claro y detallado, no apresurar el cierre.",
    }
    for perfil, n in d_perfil:
        col = PALETTE[list(dict(d_perfil).keys()).index(perfil) % len(PALETTE)] if perfil in dict(d_perfil) else "#64748b"
        desc = perfil_desc.get(perfil, "")
        html += f"""      <div class="insight" style="border-color:{col};background:{col}12;margin-bottom:10px">
        <h3 style="color:{col}">{perfil} — {n} llamadas ({pct(n,N)}%)</h3>
        <p style="color:#374151;font-size:.8rem">{desc}</p>
      </div>\n"""
    html += f"""    </div>
  </div>
</div>

<!-- ─── TAB 5 FICHAS ──────────────────────────────────────────────────── -->
<div id="s5" class="sec">
  <div class="fichas-bar">
    <input type="text" id="busq" placeholder="Buscar por cliente, objeción, coaching..." oninput="filtrar()">
    <select id="filtTipo" onchange="filtrar()">
      <option value="">Todos los fracasos</option>
      <option>Falla del Agente</option>
      <option>Falla del Agente + Otro</option>
      <option>Fricción Logística</option>
      <option>Brecha de Precio</option>
      <option>Timing — Solo Evaluando</option>
      <option>Falta de Propiedad/Decisión</option>
    </select>
    <select id="filtCal" onchange="filtrar()">
      <option value="">Todas las calificaciones C</option>
      <option>Deficiente</option>
      <option>Regular</option>
      <option>Buena</option>
    </select>
    <select id="filtPerfil" onchange="filtrar()">
      <option value="">Todos los perfiles</option>
      <option>Pragmático</option>
      <option>Timing</option>
      <option>Frustrado-Precio</option>
      <option>Desconfiado</option>
    </select>
    <span id="conteo" style="font-size:.8rem;color:#64748b;white-space:nowrap"></span>
  </div>
  <div class="fgrid" id="fgrid"></div>
</div>

<script>
const FICHAS = {fichas_data};

function show(id,el){{
  document.querySelectorAll('.sec').forEach(s=>s.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  el.classList.add('active');
}}

function filtrar(){{
  const q  = document.getElementById('busq').value.toLowerCase();
  const ft = document.getElementById('filtTipo').value;
  const fc = document.getElementById('filtCal').value;
  const fp = document.getElementById('filtPerfil').value;
  const vis = FICHAS.filter(f=>{{
    const txt = (f.cliente+f.objecion+f.coaching+f.detalle).toLowerCase();
    return (!q||txt.includes(q)) &&
           (!ft||f.tipo.includes(ft)) &&
           (!fc||f.calC===fc) &&
           (!fp||f.perfil.includes(fp));
  }});
  document.getElementById('conteo').textContent = vis.length+' llamadas';
  document.getElementById('fgrid').innerHTML = vis.map(f=>`
    <div class="fc">
      <div class="fc-head">
        <div>
          <span class="fc-name">${{f.cliente}}</span>
          <span class="fc-sub">${{f.id}} · ${{f.vehiculo}} · ${{f.tiempo}}</span>
        </div>
        <span class="badge" style="background:${{f.colorTipo}}">${{f.tipo}}</span>
      </div>
      <div class="fc-body">
        <p><strong>🧑 Perfil:</strong> ${{f.perfil}} &nbsp;|&nbsp; <strong>Tono:</strong> ${{f.tono}}</p>
        <p><strong>📢 Cita vendida como:</strong> ${{f.cita}}</p>
        <p><strong>❌ Objeción:</strong> ${{f.objecion||'—'}}</p>
        <p><strong>💥 Fracaso:</strong> ${{f.detalle||'—'}}</p>
        <p><strong>💡 Coaching:</strong> ${{f.coaching||'—'}}</p>
      </div>
      <div class="fc-foot">
        <span class="cal-badge" style="background:${{f.colorA}}">A: ${{f.calA}}</span>
        <span class="cal-badge" style="background:${{f.colorB}}">B: ${{f.calB}}</span>
        <span class="cal-badge" style="background:${{f.colorC}}">C: ${{f.calC}}</span>
      </div>
    </div>`).join('');
}}

window.onload = ()=>{{
  filtrar();
  {charts}
  new Chart(document.getElementById('chTipo2'),{{type:'bar',
    data:{{labels:{json.dumps([x[0] for x in d_tipo])},datasets:[{{
      data:{json.dumps([x[1] for x in d_tipo])},
      backgroundColor:{json.dumps([color_tipo(x[0]) for x in d_tipo])}+'cc',
      borderColor:{json.dumps([color_tipo(x[0]) for x in d_tipo])},borderWidth:1}}]}},
    options:{{indexAxis:'y',responsive:true,maintainAspectRatio:false,
      plugins:{{legend:{{display:false}}}},
      scales:{{x:{{beginAtZero:true}}}}
    }}}});
  new Chart(document.getElementById('chPerfil2'),{{type:'doughnut',
    data:{{labels:{json.dumps([x[0] for x in d_perfil])},datasets:[{{
      data:{json.dumps([x[1] for x in d_perfil])},
      backgroundColor:{json.dumps(PALETTE[:len(d_perfil)])},borderWidth:2}}]}},
    options:{{responsive:true,maintainAspectRatio:false,
      plugins:{{legend:{{position:'right',labels:{{font:{{size:10}},boxWidth:12}}}}}}
    }}}});
}};
</script>
</body></html>"""
    return html

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("Cargando datos...")
    headers, rows = cargar()
    print(f"  {len(rows)} registros cargados")

    print("Consultando Claude para síntesis...")
    client = anthropic.Anthropic()
    resumen = f"""97 llamadas sin agendamiento analizadas con Framework Comercial:

TIPIFICACIONES: Falla del Agente {pct(sum(n for t,n in [(norm_tipo_fracaso(v(r,'Tipificación de Fracaso')),1) for r in rows if v(r,'Tipificación de Fracaso')] if 'Falla' in t), len(rows))}%, Timing {pct(sum(1 for r in rows if 'Timing' in norm_tipo_fracaso(v(r,'Tipificación de Fracaso'))),len(rows))}%, Fricción Logística {pct(sum(1 for r in rows if 'Fricci' in norm_tipo_fracaso(v(r,'Tipificación de Fracaso'))),len(rows))}%, Brecha Precio {pct(sum(1 for r in rows if 'Brecha' in norm_tipo_fracaso(v(r,'Tipificación de Fracaso'))),len(rows))}%

PILAR A: Deficiente={pct(sum(1 for r in rows if 'Deficiente' in norm_cal(v(r,'A · Calificación Apertura'))),len(rows))}%, Regular={pct(sum(1 for r in rows if 'Regular' in norm_cal(v(r,'A · Calificación Apertura'))),len(rows))}%
PILAR B: Deficiente={pct(sum(1 for r in rows if 'Deficiente' in norm_cal(v(r,'B · Calificación Escucha'))),len(rows))}%
PILAR C: Deficiente={pct(sum(1 for r in rows if 'Deficiente' in norm_cal(v(r,'C · Calificación Cierre'))),len(rows))}%

TONO: Robótico/Monótono={pct(sum(1 for r in rows if 'Rob' in norm_tono(v(r,'A3 · Tono Inicial del Agente')) or 'Mon' in norm_tono(v(r,'A3 · Tono Inicial del Agente'))),len(rows))}%
SIN PREGUNTAS DESCUBRIMIENTO: {pct(sum(1 for r in rows if 'No' in norm_b1(v(r,'B1 · Preguntas de Descubrimiento'))),len(rows))}%
CITA NO VENDIDA: {pct(sum(1 for r in rows if 'No intentó' in norm_cita(v(r,'C1 · Cita Vendida Como'))),len(rows))}%
SIN SEGUNDO CIERRE: {pct(sum(1 for r in rows if 'No intentó' in norm_2cierre(v(r,'C3 · Segundo Cierre Intentado'))),len(rows))}%
ACEPTÓ PRIMER NO: {pct(sum(1 for r in rows if 'Aceptó' in norm_negativa(v(r,'C2 · Manejo Primera Negativa'))),len(rows))}%

PERFIL CLIENTE DOMINANTE: Pragmático={pct(sum(1 for r in rows if 'Pragm' in norm_perfil(v(r,'Actitud / Perfil del Cliente'))),len(rows))}%, Timing={pct(sum(1 for r in rows if 'Timing' in norm_perfil(v(r,'Actitud / Perfil del Cliente'))),len(rows))}%"""

    sintesis = sintetizar(client, resumen)
    print("  Síntesis lista.")

    print("Generando HTML...")
    html = build(headers, rows, sintesis)
    HTML_SAL.write_text(html, encoding="utf-8")
    print(f"  Panel guardado: {HTML_SAL}")
    print(f"  Tamaño: {HTML_SAL.stat().st_size//1024} KB")

if __name__ == "__main__":
    main()
