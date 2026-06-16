"""
Panel HTML — Llamadas exitosas de compra Clicar.cl
ABR / MAY / JUN hasta el 4 de junio de 2025
"""
import json, re
from pathlib import Path
from collections import Counter, defaultdict
import openpyxl
import anthropic

DATOS_DIR   = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos")
EXCEL       = DATOS_DIR / "Analisis_Llamadas_Final.xlsx"
HTML_SALIDA = DATOS_DIR / "Panel_Analisis_Llamadas.html"

# ── utilidades ────────────────────────────────────────────────────────────────

def norm(v):
    if not v: return ""
    s = str(v).strip()
    return "Sí" if s in ("Si","Sí") else s

def find(headers, frag):
    frag = frag.lower()
    for i,h in enumerate(headers):
        if h and frag in str(h).lower(): return i
    return None

def dist(rows, idx):
    vals = [norm(r[idx]) for r in rows if idx is not None and r[idx] and norm(r[idx]) not in ("","None")]
    return Counter(vals)

def pct(n, total): return round(n*100/total) if total else 0

def jl(counter, top=9):
    items = counter.most_common(top)
    return json.dumps([i[0] for i in items]), json.dumps([i[1] for i in items])

PALETTE = ["#2563eb","#22c55e","#f59e0b","#ef4444","#8b5cf6","#06b6d4","#f97316","#64748b","#84cc16"]

def bar_h(cid, title, counter, top=9, color="#2563eb"):
    L,D = jl(counter, top)
    return f"""new Chart(document.getElementById('{cid}'),{{type:'bar',
      data:{{labels:{L},datasets:[{{data:{D},backgroundColor:'{color}cc',borderColor:'{color}',borderWidth:1}}]}},
      options:{{indexAxis:'y',responsive:true,plugins:{{legend:{{display:false}},
        title:{{display:true,text:'{title}',font:{{size:13,weight:'bold'}}}}}},
        scales:{{x:{{beginAtZero:true,ticks:{{stepSize:1}}}}}}
      }}}});"""

def donut(cid, title, counter, top=7):
    items = counter.most_common(top)
    L = json.dumps([i[0] for i in items])
    D = json.dumps([i[1] for i in items])
    BG = json.dumps(PALETTE[:len(items)])
    return f"""new Chart(document.getElementById('{cid}'),{{type:'doughnut',
      data:{{labels:{L},datasets:[{{data:{D},backgroundColor:{BG},borderWidth:2}}]}},
      options:{{responsive:true,plugins:{{legend:{{position:'bottom',labels:{{font:{{size:11}}}}}},
        title:{{display:true,text:'{title}',font:{{size:13,weight:'bold'}}}}}}}}
      }});"""

def bar_v(cid, title, labels_list, datasets_list):
    L = json.dumps(labels_list)
    ds = ",".join([
        f"""{{label:'{d["label"]}',data:{json.dumps(d["data"])},
            backgroundColor:'{d["color"]}cc',borderColor:'{d["color"]}',borderWidth:1}}"""
        for d in datasets_list
    ])
    return f"""new Chart(document.getElementById('{cid}'),{{type:'bar',
      data:{{labels:{L},datasets:[{ds}]}},
      options:{{responsive:true,plugins:{{legend:{{position:'top'}},
        title:{{display:true,text:'{title}',font:{{size:13,weight:'bold'}}}}}},
        scales:{{x:{{stacked:false}},y:{{beginAtZero:true,ticks:{{stepSize:1}}}}}}
      }}}});"""

# ── carga ─────────────────────────────────────────────────────────────────────

wb = openpyxl.load_workbook(EXCEL)
ws = wb.active
headers = [cell.value for cell in ws[1]]
rows = list(ws.iter_rows(min_row=2, values_only=True))
N = len(rows)
print(f"Cargadas {N} filas.")

iA  = find(headers,"actitud")
iMo = find(headers,"motivo")
iPr = find(headers,"pregunt")
iFr = find(headers,"fricci")
iCo = find(headers,"cotiz")
iPd = find(headers,"prenda")
iSu = find(headers,"sugerencias")
iSe = find(headers,"señales")
iFi = find(headers,"ficha")
iCl = find(headers,"cliente")
iVe = find(headers,"veh")
iEs = find(headers,"estado")
iGe = find(headers,"gesti")
iIn = find(headers,"incentivo")
iAr = 0
iLi = find(headers,"licencia")
iRo = find(headers,"encargo")
iRe = find(headers,"remate")
iTi = find(headers,"titular")
iTr = find(headers,"transcri")

# distribuciones base
cA  = dist(rows,iA)
cMo = dist(rows,iMo)
cPr = dist(rows,iPr)
cFr = dist(rows,iFr)
cCo = dist(rows,iCo)
cPd = dist(rows,iPd)

# ── análisis de cruces ────────────────────────────────────────────────────────

# 1. Actitud de los que preguntaron precio vs los que no
actitud_precio_si  = Counter(norm(r[iA]) for r in rows if iPr is not None and norm(r[iPr])=="Sí" and iA is not None and r[iA])
actitud_precio_no  = Counter(norm(r[iA]) for r in rows if iPr is not None and norm(r[iPr])=="No" and iA is not None and r[iA])

# 2. Actitudes agrupadas
POSITIVAS  = {"Receptivo","Decidido","Colaborativo","Apurado","Apurada"}
NEUTRALES  = {"Analítico","Negociador","Inseguro"}
RESISTENTES= {"Escéptico","Desconfiado"}

def grupo_actitud(v):
    if v in POSITIVAS:  return "Positiva"
    if v in NEUTRALES:  return "Neutral"
    if v in RESISTENTES:return "Resistente"
    return "Otro"

grupos = Counter(grupo_actitud(norm(r[iA])) for r in rows if iA and r[iA])
n_pos = grupos.get("Positiva",0)
n_neu = grupos.get("Neutral",0)
n_res = grupos.get("Resistente",0)

# 3. Fricción por grupo de actitud
fr_por_grupo = defaultdict(Counter)
for r in rows:
    if iA and iFr and r[iA] and r[iFr]:
        g = grupo_actitud(norm(r[iA]))
        fr_por_grupo[g][norm(r[iFr])] += 1

# 4. Cotizó en otro + actitud
cotizo_actitud = Counter(norm(r[iA]) for r in rows if iCo and norm(r[iCo])=="Sí" and iA and r[iA])

# 5. Motivos vs pregunta precio
motivo_precio = defaultdict(Counter)
for r in rows:
    if iMo and iPr and r[iMo]:
        motivo_precio[norm(r[iMo])][norm(r[iPr])] += 1

# 6. Perfil "conversión rápida": actitud positiva + sin prenda + no preguntó precio
conversion_rapida = sum(
    1 for r in rows
    if iA and iPr and iPd and
       grupo_actitud(norm(r[iA]))=="Positiva" and
       norm(r[iPr])=="No" and
       norm(r[iPd])=="No"
)
perfil_tipico = pct(conversion_rapida, N)

# 7. Señales de alto interés
alto_interes = sum(1 for r in rows if iMo and norm(r[iMo]) in ("Urgencia alta","Necesita liquidez"))
cotizo_otros = cCo.get("Sí",0)
con_prenda   = sum(v for k,v in cPd.items() if k.startswith("Sí"))
sin_prenda_pct= pct(N - con_prenda, N)

# 8. Cruce: actitudes con/sin fricción
actitudes_para_grafico = [k for k,_ in cA.most_common(7)]
data_sin_friccion = [
    sum(1 for r in rows if iA and iFr and norm(r[iA])==act and norm(r[iFr]) in ("N/A","Acepta",""))
    for act in actitudes_para_grafico
]
data_con_friccion = [
    sum(1 for r in rows if iA and iFr and norm(r[iA])==act and norm(r[iFr])=="No le gusta")
    for act in actitudes_para_grafico
]

# ── síntesis Claude ───────────────────────────────────────────────────────────

print("Consultando Claude (4 llamadas)...")
client = anthropic.Anthropic()

sugerencias_txt = "\n".join(
    f"- {r[iSu]}" for r in rows
    if iSu and r[iSu] and norm(r[iSu]) not in ("Ninguna","N/A","")
)
senales_txt = "\n".join(
    f"- {r[iSe]}" for r in rows
    if iSe and r[iSe] and norm(r[iSe]) not in ("N/A","")
)[:12000]

perfiles_txt = "\n".join(
    f"actitud={norm(r[iA])}|motivo={norm(r[iMo])}|precio={norm(r[iPr])}|friccion={norm(r[iFr])}|cotizo={norm(r[iCo])}|prenda={norm(r[iPd])}"
    for r in rows
)[:6000]

stats_txt = f"""
- Total llamadas exitosas analizadas: {N} (ABR/MAY/JUN hasta 4-JUN-2025)
- Actitud positiva inicial: {n_pos} ({pct(n_pos,N)}%)
- Actitud neutral/dudosa: {n_neu} ({pct(n_neu,N)}%)
- Actitud resistente: {n_res} ({pct(n_res,N)}%)
- Preguntaron precio: {cPr.get("Sí",0)} ({pct(cPr.get("Sí",0),N)}%)
- De los que preguntaron precio, tuvieron fricción: {cFr.get("No le gusta",0)} ({pct(cFr.get("No le gusta",0), cPr.get("Sí",0))}%)
- Ya cotizaron en otro lado: {cotizo_otros} ({pct(cotizo_otros,N)}%)
- Con prenda/financiera: {con_prenda} ({pct(con_prenda,N)}%)
- Urgencia alta + Necesita liquidez: {alto_interes} ({pct(alto_interes,N)}%)
- Perfil conversión rápida (actitud positiva + sin prenda + no preguntó precio): {conversion_rapida} ({perfil_tipico}%)
"""

print("  [1/4] Factores determinantes...")
sint_determinantes = client.messages.create(model="claude-haiku-4-5", max_tokens=2000,
    messages=[{"role":"user","content":f"""Eres analista senior de ventas. Analiza estos datos estadísticos de {N} llamadas exitosas de compra de autos usados en Chile (periodo ABR/MAY/JUN 2025, hasta 4-JUN).
CONTEXTO: Queremos identificar qué factores fueron DETERMINANTES para que la llamada terminara en compra, con el fin de armar un proceso de validación rápida cuando el cliente muestra alto interés.

ESTADÍSTICAS:
{stats_txt}

SEÑALES DE COMPRA (resumen de todas las llamadas):
{senales_txt[:4000]}

Entrega:
1. Top 5-7 factores determinantes, ordenados por impacto, con explicación concreta de por qué cada uno fue clave (con datos)
2. "Perfil del cliente de conversión rápida": describe las características del cliente que convierte más rápido
3. Checklist de señales de alto interés que el agente debe detectar en los primeros 60 segundos de la llamada
4. Una conclusión ejecutiva de 2-3 oraciones
Formato: HTML con <h4>, <ul>, <li>, <strong>. Sé específico con datos y porcentajes."""}]).content[0].text.strip()

print("  [2/4] Señales de compra...")
sint_senales = client.messages.create(model="claude-haiku-4-5", max_tokens=1600,
    messages=[{"role":"user","content":f"""Analiza las señales de decisión de compra de {N} llamadas exitosas de compra de autos usados en Chile.
Identifica los 5-7 factores más recurrentes que llevaron al cliente a agendar la cita, con frecuencia aproximada.
Un insight clave para el equipo de ventas.
Formato HTML con <ul> y <li>.
SEÑALES:\n{senales_txt}"""}]).content[0].text.strip()

print("  [3/4] Sugerencias del cliente...")
sint_sugerencias = client.messages.create(model="claude-haiku-4-5", max_tokens=1800,
    messages=[{"role":"user","content":f"""Agrupa estas sugerencias de clientes en llamadas exitosas de compra de autos en Chile (máx 6 categorías).
Entrega top 3 mejoras priorizadas. Formato HTML con <h4> por categoría y <em> para ejemplos.
SUGERENCIAS:\n{sugerencias_txt}"""}]).content[0].text.strip()

print("  [4/4] Automatización...")
sint_automatizacion = client.messages.create(model="claude-haiku-4-5", max_tokens=2000,
    messages=[{"role":"user","content":f"""Analiza {N} perfiles de cliente de llamadas exitosas de compra de autos (Chile, ABR-JUN 2025).
Identifica 3-4 arquetipos con nombre, % aprox, características y flujo de contacto recomendado.
Propón 3 automatizaciones CRM/WhatsApp de alto impacto.
Formato HTML con <h4> por arquetipo.
PERFILES:\n{perfiles_txt}"""}]).content[0].text.strip()

print("  Síntesis lista.\n")

# ── fichas inspector ──────────────────────────────────────────────────────────

COLOR_ACTITUD = {
    "Receptivo":"#22c55e","Decidido":"#16a34a","Colaborativo":"#4ade80",
    "Analítico":"#3b82f6","Negociador":"#f59e0b","Inseguro":"#f97316",
    "Escéptico":"#ef4444","Desconfiado":"#dc2626","Apurado":"#8b5cf6","Apurada":"#8b5cf6"
}

fichas_html = ""
for r in rows:
    archivo  = norm(r[iAr])
    cliente  = norm(r[iCl]) if iCl else "Sin nombre"
    vehiculo = norm(r[iVe]) if iVe else ""
    estado   = norm(r[iEs]) if iEs else ""
    ficha    = norm(r[iFi]) if iFi else ""
    actitud  = norm(r[iA])  if iA  else ""
    motivo   = norm(r[iMo]) if iMo else ""
    precio_p = norm(r[iPr]) if iPr else ""
    friccion = norm(r[iFr]) if iFr else ""
    sug      = norm(r[iSu]) if iSu else ""
    gestion  = norm(r[iGe]) if iGe else ""
    prenda   = norm(r[iPd]) if iPd else ""

    color = COLOR_ACTITUD.get(actitud,"#6b7280")
    badge_precio = ("⚠️ Fricción en precio" if friccion=="No le gusta"
                    else "✓ Acepta precio" if friccion=="Acepta"
                    else "No preguntó precio")
    badge_prenda = f"<span style='color:#ef4444;font-weight:600'>⚠️ {prenda}</span>" if prenda.startswith("Sí") else ""

    fichas_html += f"""
    <div class="ficha-card" data-actitud="{actitud}" data-motivo="{motivo}" data-texto="{cliente.lower()} {vehiculo.lower()} {archivo}">
      <div class="ficha-header" style="border-left:5px solid {color}">
        <div><span class="ficha-nombre">{cliente}</span><span class="ficha-vehiculo">{vehiculo}</span></div>
        <div style="text-align:right">
          <span class="badge" style="background:{color}20;color:{color};border:1px solid {color}40">{actitud}</span>
          <span class="ficha-estado">{estado}</span>
        </div>
      </div>
      <div class="ficha-body">
        <p>{ficha}</p>
        {"<p><strong>Sugirió:</strong> <em>" + sug + "</em></p>" if sug and sug not in ("Ninguna","N/A") else ""}
        {"<p><strong>Objeciones:</strong> " + gestion + "</p>" if gestion and gestion not in ("N/A","") else ""}
      </div>
      <div class="ficha-footer">
        <span style="color:#64748b;font-size:.75rem">☎ {archivo.replace('.mp3','')}</span>
        <span class="badge-precio">{badge_precio}</span>
        {badge_prenda}
      </div>
    </div>"""

# ── charts JS ─────────────────────────────────────────────────────────────────

charts_js = (
    bar_h("ch_actitud","Actitud Inicial del Cliente",cA,10,"#2563eb") +
    donut("ch_motivo","Motivo de Venta",cMo) +
    donut("ch_precio","¿Preguntó por Precio?",cPr) +
    donut("ch_friccion","Fricción en el Precio",cFr) +
    donut("ch_cotizo","¿Ya Cotizó en Otro Lado?",cCo) +
    donut("ch_prenda","Prenda / Financiera",cPd) +
    bar_h("ch_act_r","Actitud Inicial",cA,10,"#2563eb") +
    donut("ch_mot_r","Motivo de Venta",cMo) +
    bar_v("ch_cruce_fr","Fricción en Precio por Actitud",
        actitudes_para_grafico,
        [{"label":"Sin fricción","data":data_sin_friccion,"color":"#22c55e"},
         {"label":"Con fricción","data":data_con_friccion,"color":"#ef4444"}]
    ) +
    bar_v("ch_cruce_pr","¿Preguntó Precio? por Actitud",
        actitudes_para_grafico,
        [{"label":"No preguntó","data":[sum(1 for r in rows if iA and iPr and norm(r[iA])==act and norm(r[iPr])=="No") for act in actitudes_para_grafico],"color":"#2563eb"},
         {"label":"Sí preguntó","data":[sum(1 for r in rows if iA and iPr and norm(r[iA])==act and norm(r[iPr])=="Sí") for act in actitudes_para_grafico],"color":"#f59e0b"}]
    ) +
    f"""new Chart(document.getElementById('ch_grupos'),{{type:'doughnut',
      data:{{labels:['Actitud Positiva','Actitud Neutral','Actitud Resistente'],
             datasets:[{{data:[{n_pos},{n_neu},{n_res}],
               backgroundColor:['#22c55e','#f59e0b','#ef4444'],borderWidth:2}}]}},
      options:{{responsive:true,plugins:{{legend:{{position:'bottom'}},
        title:{{display:true,text:'Grupos de Actitud',font:{{size:13,weight:'bold'}}}}}}}}
    }});"""
)

# ── HTML ──────────────────────────────────────────────────────────────────────

html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Análisis Llamadas Compra — Clicar.cl</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#f1f5f9;color:#1e293b}}
header{{background:linear-gradient(135deg,#1e3a5f,#1d4ed8);color:#fff;padding:22px 32px}}
header h1{{font-size:1.35rem;font-weight:800}}
header p{{font-size:.82rem;opacity:.85;margin-top:5px;line-height:1.5}}
.periodo-banner{{background:#fef9c3;border:1px solid #fbbf24;border-left:5px solid #f59e0b;
  padding:10px 24px;font-size:.83rem;color:#78350f;display:flex;align-items:center;gap:8px;flex-wrap:wrap}}
.tabs{{display:flex;gap:4px;background:#fff;padding:10px 20px;border-bottom:1px solid #e2e8f0;flex-wrap:wrap}}
.tab{{padding:8px 16px;border-radius:6px;cursor:pointer;font-size:.83rem;font-weight:600;
  color:#64748b;border:1px solid transparent;transition:all .15s}}
.tab:hover{{background:#f1f5f9}}.tab.active{{background:#1d4ed8;color:#fff}}
.sec{{display:none;padding:22px;max-width:1440px;margin:0 auto}}.sec.active{{display:block}}
.kpi-row{{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:10px;margin-bottom:20px}}
.kpi{{background:#fff;border-radius:10px;padding:14px;text-align:center;box-shadow:0 1px 4px #0001;border-top:3px solid transparent}}
.kpi .num{{font-size:1.9rem;font-weight:800}}.kpi .lbl{{font-size:.72rem;color:#64748b;margin-top:2px;line-height:1.3}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:16px}}
.g3{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:16px}}
.g4{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:16px}}
.card{{background:#fff;border-radius:10px;padding:18px;box-shadow:0 1px 4px #0001}}
.card h3{{font-size:.9rem;font-weight:700;margin-bottom:12px;border-bottom:2px solid #1d4ed8;padding-bottom:5px;color:#1e293b}}
.card canvas{{max-height:240px}}
.insight-box{{background:#eff6ff;border-left:5px solid #1d4ed8;border-radius:0 10px 10px 0;
  padding:14px 18px;margin-bottom:16px}}
.insight-box h3{{font-size:.9rem;font-weight:700;color:#1e40af;margin-bottom:8px}}
.insight-box p{{font-size:.83rem;line-height:1.6;color:#1e3a5f}}
.warn-box{{background:#fff7ed;border-left:5px solid #f97316;border-radius:0 10px 10px 0;
  padding:14px 18px;margin-bottom:16px}}
.warn-box h3{{font-size:.9rem;font-weight:700;color:#c2410c;margin-bottom:8px}}
.synthesis{{background:#fff;border-radius:10px;padding:20px;box-shadow:0 1px 4px #0001;line-height:1.65}}
.synthesis h4{{color:#1d4ed8;margin:14px 0 6px;font-size:.88rem}}
.synthesis ul{{padding-left:18px}}.synthesis li{{margin-bottom:4px;font-size:.85rem}}
.synthesis em{{color:#64748b;font-size:.83rem}}
.checklist{{display:grid;grid-template-columns:repeat(auto-fill,minmax(260px,1fr));gap:8px;margin-top:10px}}
.check-item{{background:#f0fdf4;border:1px solid #86efac;border-radius:8px;padding:10px 14px;font-size:.82rem;
  display:flex;align-items:flex-start;gap:8px}}
.check-item .ico{{font-size:1rem;flex-shrink:0}}
/* fichas */
.fichas-bar{{display:flex;gap:8px;margin-bottom:14px;flex-wrap:wrap;align-items:center}}
.fichas-bar input,.fichas-bar select{{padding:8px 12px;border:1px solid #cbd5e1;border-radius:6px;font-size:.83rem}}
.fichas-bar input{{flex:1;min-width:180px}}
.fgrid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(340px,1fr));gap:12px}}
.fc{{background:#fff;border-radius:10px;box-shadow:0 1px 4px #0001;overflow:hidden;font-size:.82rem}}
.fc-head{{padding:11px 13px;display:flex;justify-content:space-between;align-items:flex-start;
  background:#f8fafc;gap:8px;flex-wrap:wrap}}
.fc-name{{font-weight:700;font-size:.88rem;display:block}}
.fc-veh{{color:#64748b;font-size:.78rem}}
.fc-est{{font-size:.74rem;color:#1d4ed8;font-weight:600;display:block;text-align:right}}
.badge{{padding:2px 7px;border-radius:99px;font-size:.7rem;font-weight:600;white-space:nowrap}}
.fc-body{{padding:10px 13px;color:#374151;line-height:1.5}}.fc-body p{{margin-bottom:5px}}
.fc-foot{{padding:7px 13px;background:#f8fafc;border-top:1px solid #e2e8f0;
  display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:4px}}
.badge-precio{{font-size:.71rem;font-weight:600;color:#64748b}}
.hidden{{display:none!important}}
@media(max-width:900px){{.g2,.g3,.g4{{grid-template-columns:1fr}}}}
</style>
</head>
<body>

<header>
  <h1>Análisis de Llamadas Exitosas de Compra — Clicar.cl</h1>
  <p>{N} llamadas · Todas con cita de inspección agendada · Propósito: identificar factores determinantes de conversión para construir el proceso de validación de alto interés</p>
</header>

<div class="periodo-banner">
  📅 <strong>Período analizado: Abril / Mayo / Junio 2025 (hasta el 4 de junio)</strong>
  &nbsp;·&nbsp; ✅ Universo 100% exitoso: cada llamada resultó en una compra agendada
  &nbsp;·&nbsp; 🎯 Objetivo: entender qué gatilla la conversión rápida para automatizar el proceso de validación
</div>

<div class="tabs">
  <div class="tab active" onclick="showTab('resumen',this)">Resumen</div>
  <div class="tab" onclick="showTab('determinantes',this)">⭐ Factores Determinantes</div>
  <div class="tab" onclick="showTab('factores',this)">Perfil del Cliente</div>
  <div class="tab" onclick="showTab('voz',this)">Voz del Cliente</div>
  <div class="tab" onclick="showTab('automatizacion',this)">Automatización</div>
  <div class="tab" onclick="showTab('inspector',this)">Fichas Inspector</div>
</div>

<!-- ── RESUMEN ── -->
<div id="resumen" class="sec active">
  <div class="kpi-row">
    <div class="kpi" style="border-color:#1d4ed8"><div class="num" style="color:#1d4ed8">{N}</div><div class="lbl">Llamadas exitosas<br>ABR/MAY/JUN 2025</div></div>
    <div class="kpi" style="border-color:#22c55e"><div class="num" style="color:#22c55e">{pct(n_pos,N)}%</div><div class="lbl">Actitud positiva<br>al primer contacto</div></div>
    <div class="kpi" style="border-color:#f59e0b"><div class="num" style="color:#f59e0b">{pct(cPr.get("Sí",0),N)}%</div><div class="lbl">Preguntaron<br>por precio</div></div>
    <div class="kpi" style="border-color:#ef4444"><div class="num" style="color:#ef4444">{pct(cFr.get("No le gusta",0),N)}%</div><div class="lbl">Tuvieron fricción<br>con el precio</div></div>
    <div class="kpi" style="border-color:#8b5cf6"><div class="num" style="color:#8b5cf6">{pct(cotizo_otros,N)}%</div><div class="lbl">Ya cotizaron<br>en otro lado</div></div>
    <div class="kpi" style="border-color:#06b6d4"><div class="num" style="color:#06b6d4">{pct(con_prenda,N)}%</div><div class="lbl">Con prenda o<br>crédito vigente</div></div>
    <div class="kpi" style="border-color:#f97316"><div class="num" style="color:#f97316">{perfil_tipico}%</div><div class="lbl">Perfil conversión<br>rápida*</div></div>
  </div>
  <div class="insight-box">
    <h3>¿Qué representa este análisis?</h3>
    <p>Se analizaron <strong>{N} llamadas telefónicas</strong> del contact center de Clicar.cl realizadas entre <strong>abril y el 4 de junio de 2025</strong>, todas con resultado exitoso (cita de inspección agendada). Cada MP3 fue transcrito con IA y analizado para extraer señales de compra, actitud del cliente, fricción en precio y motivaciones. El propósito es construir un <strong>proceso de validación de alto interés</strong>: identificar rápidamente qué combinación de factores en los primeros minutos de contacto predice una conversión exitosa.</p>
    <p style="margin-top:8px;color:#1e40af;font-size:.82rem">* Perfil conversión rápida = actitud positiva + sin prenda + no preguntó precio ({conversion_rapida} de {N} clientes).</p>
  </div>
  <div class="g2">
    <div class="card"><h3>Actitud Inicial</h3><canvas id="ch_act_r"></canvas></div>
    <div class="card"><h3>Motivo de Venta</h3><canvas id="ch_mot_r"></canvas></div>
  </div>
  <div class="warn-box">
    <h3>Objetivo 1 pendiente</h3>
    <p>La evaluación de gestión del contact center se activa al incorporar las llamadas <strong>no exitosas</strong> para comparar patrones de manejo de objeciones, tiempo de conversación y tasas de cierre por agente.</p>
  </div>
</div>

<!-- ── DETERMINANTES ── -->
<div id="determinantes" class="sec">
  <div class="kpi-row">
    <div class="kpi" style="border-color:#22c55e"><div class="num" style="color:#22c55e">{pct(n_pos,N)}%</div><div class="lbl">Actitud positiva<br>(Receptivo/Decidido/Colaborativo)</div></div>
    <div class="kpi" style="border-color:#3b82f6"><div class="num" style="color:#3b82f6">{pct(N-cPr.get("Sí",0),N)}%</div><div class="lbl">Compraron sin<br>preguntar precio</div></div>
    <div class="kpi" style="border-color:#f59e0b"><div class="num" style="color:#f59e0b">{pct(cFr.get("Acepta",0), cPr.get("Sí",0))}%</div><div class="lbl">Aceptaron precio<br>entre quienes preguntaron</div></div>
    <div class="kpi" style="border-color:#8b5cf6"><div class="num" style="color:#8b5cf6">{sin_prenda_pct}%</div><div class="lbl">Sin prenda ni<br>crédito vigente</div></div>
    <div class="kpi" style="border-color:#ef4444"><div class="num" style="color:#ef4444">{pct(N-cotizo_otros,N)}%</div><div class="lbl">No habían cotizado<br>en otro lado</div></div>
  </div>
  <div class="card" style="margin-bottom:16px">
    <h3>⭐ Análisis: Factores Determinantes en la Compra</h3>
    <div class="synthesis">{sint_determinantes}</div>
  </div>
  <div class="g2">
    <div class="card"><h3>Grupos de Actitud (Positiva / Neutral / Resistente)</h3><canvas id="ch_grupos"></canvas></div>
    <div class="card"><h3>Fricción en Precio por Actitud</h3><canvas id="ch_cruce_fr"></canvas></div>
  </div>
  <div class="card" style="margin-bottom:16px">
    <h3>Pregunta de Precio por Actitud del Cliente</h3>
    <canvas id="ch_cruce_pr" style="max-height:260px"></canvas>
  </div>
</div>

<!-- ── PERFIL DEL CLIENTE ── -->
<div id="factores" class="sec">
  <div class="g3">
    <div class="card"><h3>Actitud Inicial</h3><canvas id="ch_actitud"></canvas></div>
    <div class="card"><h3>¿Preguntó por Precio?</h3><canvas id="ch_precio"></canvas></div>
    <div class="card"><h3>Fricción en el Precio</h3><canvas id="ch_friccion"></canvas></div>
  </div>
  <div class="g3">
    <div class="card"><h3>Motivo de Venta</h3><canvas id="ch_motivo"></canvas></div>
    <div class="card"><h3>¿Ya Cotizó en Otro Lado?</h3><canvas id="ch_cotizo"></canvas></div>
    <div class="card"><h3>Prenda / Financiera</h3><canvas id="ch_prenda"></canvas></div>
  </div>
  <div class="card">
    <h3>Señales de Decisión de Compra — Síntesis</h3>
    <div class="synthesis">{sint_senales}</div>
  </div>
</div>

<!-- ── VOZ DEL CLIENTE ── -->
<div id="voz" class="sec">
  <div class="insight-box">
    <h3>Voz del Cliente — Sugerencias para optimizar el proceso</h3>
    <p>Análisis de las sugerencias y comentarios recurrentes en las {N} llamadas exitosas. Estas son oportunidades de mejora directa del proceso de compra identificadas por los propios clientes que ya compraron.</p>
  </div>
  <div class="card">
    <h3>Sugerencias Agrupadas por Categoría</h3>
    <div class="synthesis">{sint_sugerencias}</div>
  </div>
</div>

<!-- ── AUTOMATIZACIÓN ── -->
<div id="automatizacion" class="sec">
  <div class="insight-box">
    <h3>Perfiles para automatización de flujos</h3>
    <p>Basado en los {N} perfiles: actitud, motivo, fricción en precio, cotización externa y estado de prenda. Los arquetipos permiten diseñar flujos diferenciados en CRM y WhatsApp para activar rápidamente cuando el cliente muestra alto interés.</p>
  </div>
  <div class="card">
    <h3>Arquetipos de Cliente y Flujos Recomendados</h3>
    <div class="synthesis">{sint_automatizacion}</div>
  </div>
</div>

<!-- ── FICHAS INSPECTOR ── -->
<div id="inspector" class="sec">
  <div class="fichas-bar">
    <input type="text" id="buscar" placeholder="Buscar por nombre, vehículo, teléfono..." oninput="filtrar()">
    <select id="filt_act" onchange="filtrar()">
      <option value="">Todas las actitudes</option>
      {"".join(f'<option value="{k}">{k} ({v})</option>' for k,v in cA.most_common())}
    </select>
    <select id="filt_mot" onchange="filtrar()">
      <option value="">Todos los motivos</option>
      {"".join(f'<option value="{k}">{k} ({v})</option>' for k,v in cMo.most_common())}
    </select>
    <span id="cnt_fichas" style="font-size:.8rem;color:#64748b;white-space:nowrap">{N} fichas</span>
  </div>
  <div class="fgrid" id="fgrid">{fichas_html}</div>
</div>

<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script>
function showTab(id,el){{
  document.querySelectorAll('.sec').forEach(s=>s.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  el.classList.add('active');
}}
function filtrar(){{
  const txt=document.getElementById('buscar').value.toLowerCase();
  const act=document.getElementById('filt_act').value;
  const mot=document.getElementById('filt_mot').value;
  let n=0;
  document.querySelectorAll('.fc').forEach(c=>{{
    const ok=(!txt||c.dataset.texto.includes(txt))&&(!act||c.dataset.actitud===act)&&(!mot||c.dataset.motivo===mot);
    c.classList.toggle('hidden',!ok);
    if(ok)n++;
  }});
  document.getElementById('cnt_fichas').textContent=n+' fichas';
}}
window.addEventListener('load',()=>{{
  {charts_js}
}});
</script>
</body></html>"""

HTML_SALIDA.write_text(html, encoding="utf-8")
print(f"Panel guardado: {HTML_SALIDA}")
