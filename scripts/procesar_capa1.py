#!/usr/bin/env python3
"""
CAPA 1 — Cliente atiende pero NO agenda cita
Transcribe MP3s y analiza con Claude para evaluar la gestión del agente:
habilidades de negociación, manejo de objeciones y momento de quiebre.

Carpeta de entrada:  LeadsCompra\llamadas_capa1\  (colocar los MP3 aquí)
Salida Excel:        LeadsCompra\datos\Analisis_Capa1_No_Agendo.xlsx

Antes de correr:
  $env:ANTHROPIC_API_KEY = (Get-ItemProperty "HKCU:\Environment" -Name "ANTHROPIC_API_KEY").ANTHROPIC_API_KEY
  $ffmpegPath = (Get-ChildItem "$env:LOCALAPPDATA\Microsoft\WinGet\Packages\" -Recurse -Filter "ffmpeg.exe").DirectoryName
  $env:PATH = "$ffmpegPath;$env:PATH"
  python scripts\procesar_capa1.py
"""

import os, json, re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import anthropic
import whisper

# ── configuración ─────────────────────────────────────────────────────────────
LLAMADAS_DIR  = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\llamadas_capa1")
DATOS_DIR     = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos")
EXCEL_SALIDA  = DATOS_DIR / "Analisis_Capa1_No_Agendo.xlsx"
PROGRESO_FILE = DATOS_DIR / ".progreso_capa1.json"

MODELO_WHISPER = "small"
MODELO_CLAUDE  = "claude-haiku-4-5"

# ── columnas ──────────────────────────────────────────────────────────────────
COLUMNAS_BASE = [
    "Nombre de Archivo",
    "Cliente",
    "Contacto Efectivo",
    "Tiempo de Conversación",
    "Actitud Inicial",
    "Vehículo / Año",
    "Patente",
    "Licencia Clase B",
    "Sin Encargo Robo",
    "Sin Remate",
    ">30 Días Titular",
    "Prenda / Financiera",
    "Gestión de Objeciones / Dudas",
    "Incentivo / Gancho Intentado",
    "Estado Final",
]
COLUMNAS_ANALISIS = [
    "Transcripción",
    "Motivo de no agendamiento",
    "Objeción principal",
    "¿El agente intentó rebatir?",
    "Calidad del manejo de objeción",
    "Momento de quiebre",
    "¿Dejó puerta abierta?",
    "Interés inicial del cliente",
    "Nivel de negociación requerido",
    "Recomendación de coaching",
]
TODAS_COLUMNAS = COLUMNAS_BASE + COLUMNAS_ANALISIS

# ── colores de cabecera ───────────────────────────────────────────────────────
COLOR_BASE     = "DC2626"  # rojo — fila de llamadas sin cierre
COLOR_ANALISIS = "7F1D1D"  # rojo oscuro — análisis de gestión

# ── prompt Claude ─────────────────────────────────────────────────────────────
PROMPT = """Eres un analista senior de contact center de compra de autos usados en Chile.
Esta llamada TERMINÓ SIN AGENDAR CITA: el cliente atendió pero no quiso reservar inspección.
Tu tarea es evaluar la gestión del agente e identificar qué falló.

TRANSCRIPCION:
{transcripcion}

Responde UNICAMENTE con JSON valido, sin texto adicional:
{{
    "cliente": "nombre completo del cliente o Desconocido",
    "contacto_efectivo": "Si",
    "tiempo_conversacion": "duracion estimada MM:SS min segun extension del dialogo",
    "actitud_inicial": "Receptivo / Inseguro / Desconfiado / Esceptico / Apurado / Molesto / Tranquilo",
    "vehiculo_anio": "Marca Modelo (AAAA) o No mencionado",
    "patente": "patente chilena si se menciona, o No mencionada",
    "licencia_clase_b": "Si / No / No menciona",
    "sin_encargo_robo": "Si / No / No menciona",
    "sin_remate": "Si / No / No menciona",
    "dias_titular": "Si / No / No menciona",
    "prenda_financiera": "Si (incluir institucion si se nombra) / No / No menciona",
    "gestion_objeciones": "descripcion de como el agente manejo las dudas y objeciones. Max 100 palabras.",
    "incentivo_gancho": "incentivo o argumento que el agente intento usar para cerrar la cita. Max 60 palabras. 'Ninguno' si no uso ninguno.",
    "estado_final": "No agendo: [motivo breve — ej: No le interesa, Precio bajo, Ya vendio, Llama despues, Corto]",
    "motivo_no_agendamiento": "razon principal por la que el cliente rechazo agendar la cita",
    "objecion_principal": "la objecion o excusa mas importante que puso el cliente. Citar textual si es posible.",
    "agente_rebatio": "Si / No / Parcialmente",
    "calidad_manejo_objecion": "Excelente / Buena / Regular / Deficiente",
    "momento_quiebre": "descripcion del momento o frase en que la llamada empezo a caer sin recuperarse. Max 80 palabras.",
    "dejo_puerta_abierta": "Si — el agente dejo abierta la posibilidad de recontacto / No / No correspondia",
    "interes_inicial": "senales de interes inicial del cliente antes de declinar, o 'Sin interes desde el inicio'. Max 80 palabras.",
    "nivel_negociacion_requerido": "Alto — el cliente necesitaba mucho convencimiento / Medio / Bajo — casi no hubo objeciones",
    "recomendacion_coaching": "recomendacion especifica y concreta para el agente basada en esta llamada. Max 100 palabras."
}}"""


# ── utilidades ────────────────────────────────────────────────────────────────

def cargar_progreso():
    if PROGRESO_FILE.exists():
        with open(PROGRESO_FILE, encoding="utf-8") as f:
            return json.load(f)
    return {}

def guardar_progreso(prog):
    PROGRESO_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(PROGRESO_FILE, "w", encoding="utf-8") as f:
        json.dump(prog, f, ensure_ascii=False, indent=2)

def transcribir(modelo_whisper, mp3_path):
    try:
        resultado = modelo_whisper.transcribe(str(mp3_path), language="es", fp16=False)
        return resultado["text"].strip()
    except Exception as e:
        return f"Error transcripcion: {e}"

def analizar_con_claude(client, transcripcion, archivo):
    texto = transcripcion[:9000]
    try:
        response = client.messages.create(
            model=MODELO_CLAUDE,
            max_tokens=1200,
            messages=[{"role": "user", "content": PROMPT.format(transcripcion=texto)}]
        )
        raw = response.content[0].text.strip()
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        return json.loads(raw)
    except Exception as e:
        print(f"  ERROR Claude {archivo}: {e}")
        return None

def crear_excel_vacio():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Analisis Capa 1"
    for j, col in enumerate(TODAS_COLUMNAS, 1):
        cell = ws.cell(row=1, column=j, value=col)
        color = COLOR_BASE if j <= len(COLUMNAS_BASE) else COLOR_ANALISIS
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"
    return wb, ws

def escribir_fila(ws, fila_num, datos_dict):
    for j, col in enumerate(TODAS_COLUMNAS, 1):
        cell = ws.cell(row=fila_num, column=j, value=datos_dict.get(col, ""))
        cell.alignment = Alignment(wrap_text=True, vertical="top")

def ajustar_anchos(ws):
    ANCHOS = {
        "Nombre de Archivo": 22, "Cliente": 22, "Transcripcion": 60,
        "Gestion de Objeciones / Dudas": 40, "Momento de quiebre": 40,
        "Recomendacion de coaching": 42, "Incentivo / Gancho Intentado": 35,
    }
    for col in ws.columns:
        header = col[0].value or ""
        max_len = max((min(len(str(c.value or "").split("\n")[0]), 60) for c in col[1:]), default=0)
        ws.column_dimensions[col[0].column_letter].width = ANCHOS.get(header, max(14, max_len + 2))

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  CAPA 1 - CLIENTE ATIENDE PERO NO AGENDA CITA")
    print("=" * 60)

    LLAMADAS_DIR.mkdir(parents=True, exist_ok=True)
    DATOS_DIR.mkdir(parents=True, exist_ok=True)

    mp3s = sorted(LLAMADAS_DIR.glob("*.mp3"))
    if not mp3s:
        print(f"\n  No hay MP3 en: {LLAMADAS_DIR}")
        print("  Coloca alli los archivos de llamadas sin agendamiento y vuelve a correr.")
        return

    print(f"\n  MP3s encontrados: {len(mp3s)}")
    print(f"  Salida: {EXCEL_SALIDA}\n")

    progreso = cargar_progreso()
    # Completo = tiene análisis Claude (más de 1 clave). Solo-transcritos = falló Claude antes.
    ya_completos  = {k for k, v in progreso.items() if len(v) > 1}
    solo_transcritos = {k: v["Transcripcion"] for k, v in progreso.items()
                        if len(v) == 1 and "Transcripcion" in v
                        and not v["Transcripcion"].startswith("Error")}
    pendientes_nuevos = [p for p in mp3s if p.name not in ya_completos and p.name not in solo_transcritos]

    print(f"  Completos: {len(ya_completos)}  |  Solo transcritos (re-analizan): {len(solo_transcritos)}  |  Nuevos: {len(pendientes_nuevos)}")

    necesita_whisper = len(pendientes_nuevos) > 0
    if necesita_whisper:
        print("\n  Cargando modelo Whisper...")
        modelo_whisper = whisper.load_model(MODELO_WHISPER)
    else:
        modelo_whisper = None
        print("  (Whisper no necesario — solo re-análisis Claude)")

    client = anthropic.Anthropic()

    # ── Re-analizar los que solo tienen transcripción ─────────────────────────
    re_lista = list(solo_transcritos.items())
    for idx, (nombre, transcripcion) in enumerate(re_lista, 1):
        print(f"  [re {idx}/{len(re_lista)}] {nombre}", end=" -> ", flush=True)
        print("Re-analizando...", end=" ", flush=True)
        datos = analizar_con_claude(client, transcripcion, nombre)
        if datos is None:
            print("FALLO (sigue sin crédito o error)")
            continue
        progreso[nombre] = {
            "Nombre de Archivo":              nombre,
            "Cliente":                        datos.get("cliente", "Desconocido"),
            "Contacto Efectivo":              datos.get("contacto_efectivo", "Si"),
            "Tiempo de Conversacion":         datos.get("tiempo_conversacion", "N/A"),
            "Actitud Inicial":                datos.get("actitud_inicial", ""),
            "Vehiculo / Anio":                datos.get("vehiculo_anio", "No mencionado"),
            "Patente":                        datos.get("patente", "No mencionada"),
            "Licencia Clase B":               datos.get("licencia_clase_b", "No menciona"),
            "Sin Encargo Robo":               datos.get("sin_encargo_robo", "No menciona"),
            "Sin Remate":                     datos.get("sin_remate", "No menciona"),
            ">30 Dias Titular":               datos.get("dias_titular", "No menciona"),
            "Prenda / Financiera":            datos.get("prenda_financiera", "No menciona"),
            "Gestion de Objeciones / Dudas":  datos.get("gestion_objeciones", ""),
            "Incentivo / Gancho Intentado":   datos.get("incentivo_gancho", ""),
            "Estado Final":                   datos.get("estado_final", "No agendo"),
            "Transcripcion":                  transcripcion,
            "Motivo de no agendamiento":      datos.get("motivo_no_agendamiento", ""),
            "Objecion principal":             datos.get("objecion_principal", ""),
            "El agente intento rebatir":      datos.get("agente_rebatio", ""),
            "Calidad del manejo de objecion": datos.get("calidad_manejo_objecion", ""),
            "Momento de quiebre":             datos.get("momento_quiebre", ""),
            "Dejo puerta abierta":            datos.get("dejo_puerta_abierta", ""),
            "Interes inicial del cliente":    datos.get("interes_inicial", ""),
            "Nivel de negociacion requerido": datos.get("nivel_negociacion_requerido", ""),
            "Recomendacion de coaching":      datos.get("recomendacion_coaching", ""),
        }
        print("OK")
        if idx % 10 == 0:
            guardar_progreso(progreso)

    guardar_progreso(progreso)

    # ── Procesar archivos nuevos (transcribir + analizar) ─────────────────────
    pendientes = pendientes_nuevos
    for idx, mp3 in enumerate(pendientes, 1):
        print(f"  [{idx}/{len(pendientes)}] {mp3.name}", end=" -> ", flush=True)

        print("Transcribiendo...", end=" ", flush=True)
        transcripcion = transcribir(modelo_whisper, mp3)

        if transcripcion.startswith("Error"):
            print(f"FALLO: {transcripcion}")
            progreso[mp3.name] = {"Transcripcion": transcripcion}
            continue

        print("Analizando...", end=" ", flush=True)
        datos = analizar_con_claude(client, transcripcion, mp3.name)
        if datos is None:
            progreso[mp3.name] = {"Transcripcion": transcripcion}
            continue

        progreso[mp3.name] = {
            "Nombre de Archivo":              mp3.name,
            "Cliente":                        datos.get("cliente", "Desconocido"),
            "Contacto Efectivo":              datos.get("contacto_efectivo", "Si"),
            "Tiempo de Conversacion":         datos.get("tiempo_conversacion", "N/A"),
            "Actitud Inicial":                datos.get("actitud_inicial", ""),
            "Vehiculo / Anio":                datos.get("vehiculo_anio", "No mencionado"),
            "Patente":                        datos.get("patente", "No mencionada"),
            "Licencia Clase B":               datos.get("licencia_clase_b", "No menciona"),
            "Sin Encargo Robo":               datos.get("sin_encargo_robo", "No menciona"),
            "Sin Remate":                     datos.get("sin_remate", "No menciona"),
            ">30 Dias Titular":               datos.get("dias_titular", "No menciona"),
            "Prenda / Financiera":            datos.get("prenda_financiera", "No menciona"),
            "Gestion de Objeciones / Dudas":  datos.get("gestion_objeciones", ""),
            "Incentivo / Gancho Intentado":   datos.get("incentivo_gancho", ""),
            "Estado Final":                   datos.get("estado_final", "No agendo"),
            "Transcripcion":                  transcripcion,
            "Motivo de no agendamiento":      datos.get("motivo_no_agendamiento", ""),
            "Objecion principal":             datos.get("objecion_principal", ""),
            "El agente intento rebatir":      datos.get("agente_rebatio", ""),
            "Calidad del manejo de objecion": datos.get("calidad_manejo_objecion", ""),
            "Momento de quiebre":             datos.get("momento_quiebre", ""),
            "Dejo puerta abierta":            datos.get("dejo_puerta_abierta", ""),
            "Interes inicial del cliente":    datos.get("interes_inicial", ""),
            "Nivel de negociacion requerido": datos.get("nivel_negociacion_requerido", ""),
            "Recomendacion de coaching":      datos.get("recomendacion_coaching", ""),
        }
        print("OK")

        if idx % 10 == 0:
            guardar_progreso(progreso)
            print(f"    [Checkpoint] {idx}/{len(pendientes)}")

    guardar_progreso(progreso)

    # Generar Excel
    print("\n  Generando Excel...")
    wb, ws = crear_excel_vacio()

    # Mapa de clave interna a nombre de columna en el Excel
    MAPA_COLUMNAS = {
        "Nombre de Archivo":              "Nombre de Archivo",
        "Cliente":                        "Cliente",
        "Contacto Efectivo":              "Contacto Efectivo",
        "Tiempo de Conversacion":         "Tiempo de Conversación",
        "Actitud Inicial":                "Actitud Inicial",
        "Vehiculo / Anio":                "Vehículo / Año",
        "Patente":                        "Patente",
        "Licencia Clase B":               "Licencia Clase B",
        "Sin Encargo Robo":               "Sin Encargo Robo",
        "Sin Remate":                     "Sin Remate",
        ">30 Dias Titular":               ">30 Días Titular",
        "Prenda / Financiera":            "Prenda / Financiera",
        "Gestion de Objeciones / Dudas":  "Gestión de Objeciones / Dudas",
        "Incentivo / Gancho Intentado":   "Incentivo / Gancho Intentado",
        "Estado Final":                   "Estado Final",
        "Transcripcion":                  "Transcripción",
        "Motivo de no agendamiento":      "Motivo de no agendamiento",
        "Objecion principal":             "Objeción principal",
        "El agente intento rebatir":      "¿El agente intentó rebatir?",
        "Calidad del manejo de objecion": "Calidad del manejo de objeción",
        "Momento de quiebre":             "Momento de quiebre",
        "Dejo puerta abierta":            "¿Dejó puerta abierta?",
        "Interes inicial del cliente":    "Interés inicial del cliente",
        "Nivel de negociacion requerido": "Nivel de negociación requerido",
        "Recomendacion de coaching":      "Recomendación de coaching",
    }

    for fila_num, (archivo, datos) in enumerate(progreso.items(), 2):
        row_data = {MAPA_COLUMNAS.get(k, k): v for k, v in datos.items()}
        row_data["Nombre de Archivo"] = archivo
        escribir_fila(ws, fila_num, row_data)

    ajustar_anchos(ws)
    wb.save(EXCEL_SALIDA)

    total = len(progreso)
    print(f"\n{'=' * 60}")
    print(f"  COMPLETADO")
    print(f"  Excel: {EXCEL_SALIDA}")
    print(f"  Total registros: {total}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
