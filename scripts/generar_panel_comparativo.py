"""
Panel Comparativo de las 3 Capas de Conversión — Clicar.cl
  Capa 1: cliente atiende pero NO agenda
  Capa 2: cliente agenda pero NO compra
  Capa 3: cliente agenda Y compra (casos de éxito)

Corre con cualquier combinación de datos disponibles.
Salida: datos/Panel_Comparativo_Capas.html
"""
import json, re
from pathlib import Path
from collections import Counter, defaultdict
import openpyxl
import anthropic

DATOS_DIR  = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos")
HTML_SAL   = DATOS_DIR / "Panel_Comparativo_Capas.html"

EXCEL_C1   = DATOS_DIR / "Analisis_Capa1_No_Agendo.xlsx"
EXCEL_C2   = DATOS_DIR / "Analisis_Capa2_No_Compro.xlsx"
EXCEL_C3   = DATOS_DIR / "Analisis_Llamadas_Final.xlsx"

# paleta por capa
PAL = {1: "#dc2626", 2: "#d97706", 3: "#2563eb"}
PAL_LIGHT = {1: "#fef2f2", 2: "#fffbeb", 3: "#eff6ff"}
PAL_TEXT  = {1: "#991b1b", 2: "#92400e", 3: "#1e40af"}
LABEL = {1: "Capa 1 — No agendó", 2: "Capa 2 — Agendó, no compró", 3: "Capa 3 — Compra exitosa"}
ICON  = {1: "❌", 2: "⚠️", 3: "✅"}

# paleta general para gráficos
PALETTE = ["#2563eb","#22c55e","#f59e0b","#ef4444","#8b5cf6","#06b6d4","#f97316","#64748b","#84cc16","#ec4899"]

# ── carga genérica ────────────────────────────────────────────────────────────

def find_col(headers, frag):
    frag = frag.lower()
    for i, h in enumerate(headers):
        if h and frag in str(h).lower():
            return i
    return None

def norm(v):
    if v is None: return ""
    s = str(v).strip()
    return "Sí" if s in ("Si","Sí") else s

def cargar_excel(path):
    """Retorna (headers, rows) o None si no existe."""
    if not path.exists():
        return None
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    rows = [r for r in rows if any(r)]
    return headers, rows

def dist(rows, idx, top=10):
    if idx is None: return Counter()
    vals = [norm(r[idx]) for r in rows if r[idx] and norm(r[idx]) not in ("","None","N/A","No menciona","No mencionado")]
    return Counter(vals)

def pct(n, total): return round(n*100/total) if total else 0

# ── JS helpers ────────────────────────────────────────────────────────────────

def jl(counter, top=9):
    items = counter.most_common(top)
    return json.dumps([i[0] for i in items]), json.dumps([i[1] for i in items])

def bar_h(cid, title, counter, top=9, color="#2563eb"):
    L, D = jl(counter, top)
    return f"""new Chart(document.getElementById('{cid}'),{{type:'bar',
      data:{{labels:{L},datasets:[{{data:{D},backgroundColor:'{color}cc',borderColor:'{color}',borderWidth:1}}]}},
      options:{{indexAxis:'y',responsive:true,plugins:{{legend:{{display:false}},
        title:{{display:true,text:'{title}',font:{{size:12,weight:'bold'}}}}}},
        scales:{{x:{{beginAtZero:true,ticks:{{stepSize:1}}}}}}
      }}}});"""

def donut(cid, title, counter, top=7):
    items = counter.most_common(top)
    L = json.dumps([i[0] for i in items])
    D = json.dumps([i[1] for i in items])
    BG = json.dumps(PALETTE[:len(items)])
    return f"""new Chart(document.getElementById('{cid}'),{{type:'doughnut',
      data:{{labels:{L},datasets:[{{data:{D},backgroundColor:{BG},borderWidth:2}}]}},
      options:{{responsive:true,plugins:{{legend:{{position:'bottom',labels:{{font:{{size:10}}}}}},
        title:{{display:true,text:'{title}',font:{{size:12,weight:'bold'}}}}}}}}
      }});"""

def bar_v_multi(cid, title, labels, datasets):
    L = json.dumps(labels)
    ds = ",".join([
        f"""{{label:'{d["label"]}',data:{json.dumps(d["data"])},
            backgroundColor:'{d["color"]}bb',borderColor:'{d["color"]}',borderWidth:1}}"""
        for d in datasets
    ])
    return f"""new Chart(document.getElementById('{cid}'),{{type:'bar',
      data:{{labels:{L},datasets:[{ds}]}},
      options:{{responsive:true,plugins:{{legend:{{position:'top'}},
        title:{{display:true,text:'{title}',font:{{size:12,weight:'bold'}}}}}},
        scales:{{x:{{ticks:{{font:{{size:10}}}}}},y:{{beginAtZero:true,ticks:{{stepSize:1}}}}}}
      }}}});"""

# ── estado pendiente ──────────────────────────────────────────────────────────

def pending_section(capa, excel_name, script_name, carpeta):
    color = PAL[capa]
    return f"""
    <div style="text-align:center;padding:60px 20px;background:#f8fafc;border-radius:12px;border:2px dashed {color}40;margin:16px 0">
      <div style="font-size:3rem;margin-bottom:12px">⏳</div>
      <h3 style="color:{color};font-size:1.1rem;margin-bottom:8px">{LABEL[capa]} — Sin datos cargados</h3>
      <p style="color:#64748b;font-size:.85rem;max-width:480px;margin:0 auto 16px">
        Para activar este análisis, coloca los MP3 correspondientes en la carpeta indicada y ejecuta el script.
      </p>
      <div style="background:#fff;border:1px solid #e2e8f0;border-radius:8px;padding:14px 20px;display:inline-block;text-align:left;font-size:.82rem;color:#374151">
        <div style="margin-bottom:6px"><strong>Carpeta:</strong> <code>LeadsCompra\\{carpeta}\\</code></div>
        <div style="margin-bottom:6px"><strong>Excel de salida:</strong> <code>datos\\{excel_name}</code></div>
        <div><strong>Script:</strong> <code>python scripts\\{script_name}</code></div>
      </div>
      <p style="color:#94a3b8;font-size:.78rem;margin-top:14px">
        Después vuelve a correr <code>python scripts\\generar_panel_comparativo.py</code>
      </p>
    </div>"""

# ── síntesis Claude ───────────────────────────────────────────────────────────

def sintetizar(client, prompt_text):
    try:
        r = client.messages.create(model="claude-haiku-4-5", max_tokens=1800,
            messages=[{"role":"user","content":prompt_text}])
        return r.content[0].text.strip()
    except Exception as e:
        return f"<p style='color:#ef4444'>Error síntesis: {e}</p>"

# ── cálculo grupos actitud ────────────────────────────────────────────────────

POSITIVAS   = {"Receptivo","Decidido","Colaborativo","Apurado","Apurada","Curioso"}
NEUTRALES   = {"Analítico","Negociador","Inseguro","Tranquilo"}
RESISTENTES = {"Escéptico","Desconfiado","Molesto"}

def grupo_actitud(v):
    if v in POSITIVAS:   return "Positiva"
    if v in NEUTRALES:   return "Neutral"
    if v in RESISTENTES: return "Resistente"
    return "Otro"

def grupos_pct(rows, idx_a):
    if idx_a is None or not rows: return (0,0,0,0)
    n = len(rows)
    pos = sum(1 for r in rows if grupo_actitud(norm(r[idx_a]))=="Positiva")
    neu = sum(1 for r in rows if grupo_actitud(norm(r[idx_a]))=="Neutral")
    res = sum(1 for r in rows if grupo_actitud(norm(r[idx_a]))=="Resistente")
    return pct(pos,n), pct(neu,n), pct(res,n), n

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("Cargando datos...")
    datos = {}
    for capa, path in [(1,EXCEL_C1),(2,EXCEL_C2),(3,EXCEL_C3)]:
        r = cargar_excel(path)
        if r:
            headers, rows = r
            datos[capa] = {"headers":headers,"rows":rows,"n":len(rows)}
            print(f"  Capa {capa}: {len(rows)} registros OK  [{path.name}]")
        else:
            print(f"  Capa {capa}: sin datos  [{path.name}]")

    capas_ok = sorted(datos.keys())
    n_total  = sum(d["n"] for d in datos.values())

    # índices por capa
    idx = {}
    for capa, d in datos.items():
        h = d["headers"]
        idx[capa] = {
            "a": find_col(h,"actitud"),
            "mo":find_col(h,"motivo"),
            "pr":find_col(h,"pregunt"),
            "fr":find_col(h,"fricci"),
            "co":find_col(h,"cotiz"),
            "pd":find_col(h,"prenda"),
            "se":find_col(h,"senales") or find_col(h,"señales"),
            "su":find_col(h,"suger"),
            # capa 1 específico
            "ob":find_col(h,"objecion") or find_col(h,"objeción"),
            "ca":find_col(h,"calidad"),
            "nv":find_col(h,"nivel"),
            "mo_na":find_col(h,"no agend"),
            # capa 2 específico
            "br":find_col(h,"brecha"),
            "ns":find_col(h,"no show"),
            "rc":find_col(h,"riesgo"),
            "ep":find_col(h,"expectativa"),
        }

    # distribuciones
    dists = {}
    for capa, d in datos.items():
        rows = d["rows"]
        ix   = idx[capa]
        dists[capa] = {
            "a":  dist(rows, ix["a"]),
            "mo": dist(rows, ix["mo"]),
            "pr": dist(rows, ix["pr"]),
            "fr": dist(rows, ix["fr"]),
            "co": dist(rows, ix["co"]),
            "pd": dist(rows, ix["pd"]),
        }

    # grupos de actitud
    gp = {}
    for capa, d in datos.items():
        gp[capa] = grupos_pct(d["rows"], idx[capa]["a"])

    # síntesis Claude
    print("Consultando Claude...")
    client = anthropic.Anthropic()

    sint = {}
    # capa 3 siempre disponible
    if 3 in datos:
        d3 = datos[3]; ix3 = idx[3]; dists3 = dists[3]
        senales_txt = "\n".join(
            f"- {r[ix3['se']]}" for r in d3["rows"]
            if ix3["se"] and r[ix3["se"]] and norm(r[ix3["se"]]) not in ("N/A","")
        )[:6000]
        print("  [1] Síntesis Capa 3...")
        sint[3] = sintetizar(client, f"""Eres analista de ventas de autos usados Chile.
Analiza {d3['n']} llamadas exitosas (ABR-JUN 2025, cita + compra confirmada).
Estadísticas:
- Actitud positiva: {gp[3][0]}% | Neutral: {gp[3][1]}% | Resistente: {gp[3][2]}%
- Preguntaron precio: {dists3["pr"].get("Sí",0)} ({pct(dists3["pr"].get("Sí",0),d3["n"])}%)
- Con fricción precio: {dists3["fr"].get("No le gusta",0)} ({pct(dists3["fr"].get("No le gusta",0),d3["n"])}%)
- Cotizaron en otro lado: {dists3["co"].get("Sí",0)} ({pct(dists3["co"].get("Sí",0),d3["n"])}%)
- Con prenda: {sum(v for k,v in dists3["pd"].items() if k.startswith("Sí"))} ({pct(sum(v for k,v in dists3["pd"].items() if k.startswith("Sí")),d3["n"])}%)
Señales de compra detectadas:
{senales_txt}
Entrega: top 6 factores de conversión con datos, perfil del cliente que cierra rápido, y 3 señales de alto interés a detectar en <60s.
Formato HTML con <h4>, <ul>, <li>, <strong>.""")

    if 1 in datos:
        d1 = datos[1]; ix1 = idx[1]
        cCalidad = dist(d1["rows"], ix1["ca"])
        cNivel   = dist(d1["rows"], ix1["nv"])
        cObjecion= dist(d1["rows"], ix1["ob"])
        cNoAg    = dist(d1["rows"], ix1["mo_na"])
        print("  [2] Síntesis Capa 1...")
        sint[1] = sintetizar(client, f"""Analiza {d1['n']} llamadas donde el cliente atendió pero NO agendó cita (clicar.cl).
Distribución calidad manejo objeción: {dict(cCalidad.most_common(5))}
Distribución nivel negociación requerido: {dict(cNivel.most_common())}
Motivos de no agendamiento más frecuentes: {dict(cNoAg.most_common(8))}
Objeciones principales: {dict(cObjecion.most_common(8))}
Actitud inicial: {dict(dists[1]["a"].most_common(7))}
Entrega: 1) los 3 principales errores de gestión detectados, 2) las 3 habilidades que más faltaron, 3) recomendaciones de entrenamiento concretas.
Formato HTML con <h4>, <ul>, <li>.""")

    if 2 in datos:
        d2 = datos[2]; ix2 = idx[2]
        cBrecha = dist(d2["rows"], ix2["br"])
        cNS     = dist(d2["rows"], ix2["ns"])
        cRC     = dist(d2["rows"], ix2["rc"])
        print("  [3] Síntesis Capa 2...")
        sint[2] = sintetizar(client, f"""Analiza {d2['n']} llamadas donde el cliente agendó cita pero NO compró (clicar.cl).
Brecha precio-expectativa: {dict(cBrecha.most_common())}
Riesgo de no show: {dict(cNS.most_common())}
Riesgos comerciales: {dict(cRC.most_common(8))}
Actitud inicial: {dict(dists[2]["a"].most_common(7))}
Entrega: 1) principales causas de la no compra, 2) señales de riesgo que se podían detectar en la llamada, 3) recomendaciones para el inspector para rescatar estos casos.
Formato HTML con <h4>, <ul>, <li>.""")

    if len(capas_ok) >= 2:
        print("  [4] Síntesis comparativa...")
        resumen_capas = "\n".join([
            f"Capa {c}: n={datos[c]['n']}, actitud positiva={gp[c][0]}%, neutral={gp[c][1]}%, resistente={gp[c][2]}%"
            for c in capas_ok
        ])
        sint["comp"] = sintetizar(client, f"""Compara {len(capas_ok)} capas de llamadas de compra de autos usados (clicar.cl):
{resumen_capas}
¿Qué diferencia al cliente que compra (capa 3) del que no agenda (capa 1) y del que agenda pero no compra (capa 2)?
Entrega: 3-4 diferencias clave con datos, y un protocolo de acción rápida para cuando el cliente muestre alto interés.
Formato HTML con <h4>, <ul>, <li>.""")
    else:
        sint["comp"] = None

    print("  Síntesis lista.\n")

    # ── construcción de la comparación de actitudes ───────────────────────────
    top_actitudes = []
    for c in capas_ok:
        top_actitudes += list(dists[c]["a"].keys())
    top_actitudes = list(dict.fromkeys(  # orden de primera aparición
        k for k,_ in Counter(top_actitudes).most_common(8)
    ))

    charts_js = ""

    # embudo - horizontal bar manual
    funnel_data = [datos[c]["n"] if c in datos else 0 for c in [1,2,3]]
    funnel_labels = [
        f"C1 — Atiende, no agenda ({funnel_data[0]})" if funnel_data[0] else "C1 — Sin datos",
        f"C2 — Agenda, no compra  ({funnel_data[1]})" if funnel_data[1] else "C2 — Sin datos",
        f"C3 — Compra exitosa      ({funnel_data[2]})" if funnel_data[2] else "C3 — Sin datos",
    ]
    charts_js += f"""new Chart(document.getElementById('ch_embudo'),{{type:'bar',
      data:{{labels:{json.dumps(funnel_labels)},
             datasets:[{{data:{json.dumps(funnel_data)},
               backgroundColor:['{PAL[1]}bb','{PAL[2]}bb','{PAL[3]}bb'],
               borderColor:['{PAL[1]}','{PAL[2]}','{PAL[3]}'],borderWidth:2}}]}},
      options:{{indexAxis:'y',responsive:true,plugins:{{legend:{{display:false}},
        title:{{display:true,text:'Volumen por capa de conversión',font:{{size:13,weight:'bold'}}}}}},
        scales:{{x:{{beginAtZero:true}}}}
      }}}});"""

    # grupos actitud por capa
    if len(capas_ok) >= 1:
        cap_labels = [LABEL[c] for c in capas_ok]
        charts_js += bar_v_multi("ch_grupos_act","Grupos de Actitud por Capa (%)",
            cap_labels,
            [{"label":"Positiva","data":[gp[c][0] for c in capas_ok],"color":"#22c55e"},
             {"label":"Neutral", "data":[gp[c][1] for c in capas_ok],"color":"#f59e0b"},
             {"label":"Resistente","data":[gp[c][2] for c in capas_ok],"color":"#ef4444"}]
        )

    # actitudes comparadas entre capas (absolutos)
    if len(capas_ok) >= 2:
        act_datasets = [
            {"label":LABEL[c],"data":[dists[c]["a"].get(a,0) for a in top_actitudes],"color":PAL[c]}
            for c in capas_ok
        ]
        charts_js += bar_v_multi("ch_act_comp","Actitud Inicial — Comparación entre Capas",
            top_actitudes, act_datasets)

        # precio comparado
        precio_labels = ["Preguntó precio","No preguntó precio"]
        precio_ds = []
        for c in capas_ok:
            si = dists[c]["pr"].get("Sí",0)
            no = dists[c]["pr"].get("No",0)
            precio_ds.append({"label":LABEL[c],"data":[si,no],"color":PAL[c]})
        charts_js += bar_v_multi("ch_precio_comp","Pregunta de Precio — Comparación",
            precio_labels, precio_ds)

        # friccion comparada
        fr_labels = ["Acepta","No le gusta","N/A"]
        fr_ds = []
        for c in capas_ok:
            fr_ds.append({"label":LABEL[c],
                "data":[dists[c]["fr"].get(k,0) for k in fr_labels],"color":PAL[c]})
        charts_js += bar_v_multi("ch_fr_comp","Fricción en Precio — Comparación",
            fr_labels, fr_ds)

    # gráficos individuales capa 3
    if 3 in datos:
        charts_js += bar_h("ch_c3_act","Actitud Inicial",dists[3]["a"],10,PAL[3])
        charts_js += donut("ch_c3_mot","Motivo de Venta",dists[3]["mo"])
        charts_js += donut("ch_c3_pr","Preguntó Precio",dists[3]["pr"])
        charts_js += donut("ch_c3_fr","Fricción Precio",dists[3]["fr"])
        charts_js += donut("ch_c3_co","Cotizó en Otro Lado",dists[3]["co"])
        charts_js += donut("ch_c3_pd","Prenda / Financiera",dists[3]["pd"])

    # gráficos individuales capa 1
    if 1 in datos:
        charts_js += bar_h("ch_c1_act","Actitud Inicial",dists[1]["a"],10,PAL[1])
        charts_js += donut("ch_c1_ca","Calidad Manejo Objeción",dist(datos[1]["rows"],idx[1]["ca"]))
        charts_js += donut("ch_c1_nv","Nivel de Negociación Requerido",dist(datos[1]["rows"],idx[1]["nv"]))
        charts_js += bar_h("ch_c1_ob","Objeción Principal",dist(datos[1]["rows"],idx[1]["ob"]),8,PAL[1])

    # gráficos individuales capa 2
    if 2 in datos:
        charts_js += bar_h("ch_c2_act","Actitud Inicial",dists[2]["a"],10,PAL[2])
        charts_js += donut("ch_c2_br","Brecha Precio-Expectativa",dist(datos[2]["rows"],idx[2]["br"]))
        charts_js += donut("ch_c2_ns","Riesgo de No Show",dist(datos[2]["rows"],idx[2]["ns"]))
        charts_js += bar_h("ch_c2_rc","Riesgo Comercial",dist(datos[2]["rows"],idx[2]["rc"]),8,PAL[2])

    # ── estado de datos para la barra superior ────────────────────────────────
    def estado_badge(capa):
        if capa in datos:
            n = datos[capa]["n"]
            bg, txt = PAL_LIGHT[capa], PAL_TEXT[capa]
            return f"""<span style="background:{bg};color:{txt};border:1px solid {PAL[capa]}40;
              padding:4px 10px;border-radius:99px;font-size:.78rem;font-weight:600">
              {ICON[capa]} Capa {capa}: {n} registros</span>"""
        return f"""<span style="background:#f1f5f9;color:#94a3b8;border:1px solid #e2e8f0;
          padding:4px 10px;border-radius:99px;font-size:.78rem;font-weight:600">
          ⏳ Capa {capa}: sin datos</span>"""

    def tab_label(capa):
        status = "✅" if capa in datos else "⏳"
        names = {1:"No agendó",2:"No compró",3:"Compra exitosa"}
        n = f" ({datos[capa]['n']})" if capa in datos else ""
        return f"{status} Capa {capa} — {names[capa]}{n}"

    def card_kpi(num, lbl, color):
        return f"""<div class="kpi" style="border-color:{color}">
          <div class="num" style="color:{color}">{num}</div>
          <div class="lbl">{lbl}</div></div>"""

    # totales para el embudo
    n1 = datos[1]["n"] if 1 in datos else "?"
    n2 = datos[2]["n"] if 2 in datos else "?"
    n3 = datos[3]["n"] if 3 in datos else "?"

    # tasa conversión (solo si tenemos ambos)
    if 1 in datos and 3 in datos:
        total_atienden = datos[1]["n"] + (datos[2]["n"] if 2 in datos else 0) + datos[3]["n"]
        tasa_ag = pct((datos[2]["n"] if 2 in datos else 0) + datos[3]["n"], total_atienden)
        tasa_comp = pct(datos[3]["n"], (datos[2]["n"] if 2 in datos else 0) + datos[3]["n"])
    else:
        tasa_ag = "?"
        tasa_comp = "?"

    # ── HTML ──────────────────────────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Panel Comparativo de Conversión — Clicar.cl</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Segoe UI',system-ui,sans-serif;background:#f1f5f9;color:#1e293b}}
header{{background:linear-gradient(135deg,#1e3a5f,#1d4ed8);color:#fff;padding:20px 28px}}
header h1{{font-size:1.25rem;font-weight:800}}header p{{font-size:.8rem;opacity:.85;margin-top:4px}}
.estado-bar{{background:#fff;padding:10px 24px;border-bottom:1px solid #e2e8f0;
  display:flex;gap:8px;flex-wrap:wrap;align-items:center}}
.estado-bar span.lbl{{font-size:.78rem;color:#64748b;margin-right:4px}}
.tabs{{display:flex;gap:4px;background:#fff;padding:8px 18px;border-bottom:1px solid #e2e8f0;
  overflow-x:auto;white-space:nowrap}}
.tab{{padding:7px 14px;border-radius:6px;cursor:pointer;font-size:.8rem;font-weight:600;
  color:#64748b;border:1px solid transparent;white-space:nowrap}}
.tab:hover{{background:#f1f5f9}}.tab.active{{background:#1d4ed8;color:#fff}}
.sec{{display:none;padding:20px;max-width:1440px;margin:0 auto}}.sec.active{{display:block}}
.kpi-row{{display:grid;grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px;margin-bottom:18px}}
.kpi{{background:#fff;border-radius:10px;padding:12px;text-align:center;box-shadow:0 1px 4px #0001;
  border-top:3px solid transparent}}
.kpi .num{{font-size:1.8rem;font-weight:800}}.kpi .lbl{{font-size:.72rem;color:#64748b;margin-top:2px;line-height:1.3}}
.g2{{display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px}}
.g3{{display:grid;grid-template-columns:repeat(3,1fr);gap:14px;margin-bottom:14px}}
.g4{{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:14px}}
.card{{background:#fff;border-radius:10px;padding:18px;box-shadow:0 1px 4px #0001;margin-bottom:14px}}
.card h3{{font-size:.88rem;font-weight:700;margin-bottom:12px;color:#1e293b;
  border-bottom:2px solid #1d4ed8;padding-bottom:5px}}
.card canvas{{max-height:230px}}
.insight{{border-left:5px solid;border-radius:0 10px 10px 0;padding:14px 18px;margin-bottom:14px}}
.insight h3{{font-size:.88rem;font-weight:700;margin-bottom:8px}}
.insight p{{font-size:.83rem;line-height:1.6}}
.synthesis{{background:#fff;border-radius:10px;padding:18px;box-shadow:0 1px 4px #0001;
  line-height:1.65;font-size:.85rem}}
.synthesis h4{{color:#1d4ed8;margin:12px 0 5px;font-size:.88rem}}
.synthesis ul{{padding-left:16px}}.synthesis li{{margin-bottom:4px}}
.capa-header{{padding:12px 16px;border-radius:10px;margin-bottom:16px;display:flex;
  align-items:center;gap:12px}}
.capa-header .num{{font-size:2.5rem;font-weight:900;opacity:.15}}
.capa-header .info h2{{font-size:1rem;font-weight:700}}.capa-header .info p{{font-size:.8rem;opacity:.85}}
@media(max-width:900px){{.g2,.g3,.g4{{grid-template-columns:1fr}}}}
</style>
</head>
<body>

<header>
  <h1>Panel Comparativo de Conversión — Clicar.cl</h1>
  <p>Análisis de las 3 capas del embudo: contacto / agendamiento / compra · ABR-JUN 2025</p>
</header>

<div class="estado-bar">
  <span class="lbl">Estado de datos:</span>
  {estado_badge(1)}{estado_badge(2)}{estado_badge(3)}
</div>

<div class="tabs">
  <div class="tab active" onclick="showTab('embudo',this)">🔽 Embudo</div>
  <div class="tab" onclick="showTab('c1',this)">{tab_label(1)}</div>
  <div class="tab" onclick="showTab('c2',this)">{tab_label(2)}</div>
  <div class="tab" onclick="showTab('c3',this)">{tab_label(3)}</div>
  <div class="tab" onclick="showTab('comp',this)">📊 Comparación</div>
</div>

<!-- ── EMBUDO ── -->
<div id="embudo" class="sec active">
  <div class="kpi-row">
    {card_kpi(n1,"Atendieron<br>no agendaron",PAL[1])}
    {card_kpi(n2,"Agendaron<br>no compraron",PAL[2])}
    {card_kpi(n3,"Agendaron<br>y compraron",PAL[3])}
    {card_kpi(f"{tasa_ag}%" if tasa_ag!="?" else "—","Tasa de<br>agendamiento","#8b5cf6")}
    {card_kpi(f"{tasa_comp}%" if tasa_comp!="?" else "—","Tasa de cierre<br>desde cita","#22c55e")}
    {card_kpi(n_total,"Total llamadas<br>analizadas","#64748b")}
  </div>
  <div class="card">
    <h3>Distribución por capa</h3>
    <canvas id="ch_embudo" style="max-height:180px"></canvas>
  </div>
  {"" if len(capas_ok)<2 else f'''<div class="g2">
    <div class="card"><h3>Grupos de Actitud por Capa (%)</h3><canvas id="ch_grupos_act"></canvas></div>
    <div class="card">
      <h3>Interpretación del embudo</h3>
      <div class="synthesis">{sint.get("comp","") or
        "<p>Análisis comparativo disponible cuando estén cargadas al menos 2 capas.</p>"}</div>
    </div>
  </div>'''}
  <div class="insight" style="background:#f0fdf4;border-color:#22c55e">
    <h3 style="color:#166534">¿Qué representa cada capa?</h3>
    <p>
      <strong style="color:{PAL[1]}">Capa 1 (rojo)</strong> — El agente logró que el cliente atendiera, pero no cerró la cita.
      Permite evaluar habilidades de negociación y manejo de objeciones del contact center.<br>
      <strong style="color:{PAL[2]}">Capa 2 (naranja)</strong> — El agente agendó la cita, pero el cliente no compró.
      Permite analizar qué fricción comercial (precio, desconfianza, prenda) ocurrió en la visita y si había señales en la llamada.<br>
      <strong style="color:{PAL[3]}">Capa 3 (azul)</strong> — Conversiones exitosas.
      Base de comportamiento ideal para construir el proceso de validación de alto interés.
    </p>
  </div>
</div>

<!-- ── CAPA 1 ── -->
<div id="c1" class="sec">
  <div class="capa-header" style="background:{PAL_LIGHT[1]};color:{PAL_TEXT[1]}">
    <div class="num">1</div>
    <div class="info">
      <h2>{ICON[1]} Capa 1 — Cliente atiende pero NO agenda cita</h2>
      <p>Objetivo: evaluar gestión del contact center · Habilidades de negociación y manejo de objeciones</p>
    </div>
  </div>
  {pending_section(1,'Analisis_Capa1_No_Agendo.xlsx','procesar_capa1.py','llamadas_capa1') if 1 not in datos else f'''
  <div class="kpi-row">
    {card_kpi(datos[1]["n"],"Llamadas<br>sin agendamiento",PAL[1])}
    {card_kpi(gp[1][0],"% Actitud<br>positiva inicial","#22c55e")}
    {card_kpi(gp[1][2],"% Actitud<br>resistente","#ef4444")}
    {card_kpi(pct(dist(datos[1]["rows"],idx[1]["ca"]).get("Deficiente",0)+dist(datos[1]["rows"],idx[1]["ca"]).get("Regular",0),datos[1]["n"]),"% Manejo objeción<br>Regular o Deficiente","#f59e0b")}
  </div>
  <div class="g2">
    <div class="card"><h3>Actitud Inicial</h3><canvas id="ch_c1_act"></canvas></div>
    <div class="card"><h3>Calidad del Manejo de Objeción</h3><canvas id="ch_c1_ca"></canvas></div>
  </div>
  <div class="g2">
    <div class="card"><h3>Nivel de Negociación Requerido</h3><canvas id="ch_c1_nv"></canvas></div>
    <div class="card"><h3>Objeción Principal</h3><canvas id="ch_c1_ob"></canvas></div>
  </div>
  <div class="card">
    <h3>Análisis — Gestión del Contact Center</h3>
    <div class="synthesis">{sint.get(1,"")}</div>
  </div>'''}
</div>

<!-- ── CAPA 2 ── -->
<div id="c2" class="sec">
  <div class="capa-header" style="background:{PAL_LIGHT[2]};color:{PAL_TEXT[2]}">
    <div class="num">2</div>
    <div class="info">
      <h2>{ICON[2]} Capa 2 — Cliente agenda cita pero NO compra</h2>
      <p>Objetivo: identificar fricción comercial · Señales de riesgo detectables en la llamada</p>
    </div>
  </div>
  {pending_section(2,'Analisis_Capa2_No_Compro.xlsx','procesar_capa2.py','llamadas_capa2') if 2 not in datos else f'''
  <div class="kpi-row">
    {card_kpi(datos[2]["n"],"Llamadas con cita<br>sin compra",PAL[2])}
    {card_kpi(gp[2][0],"% Actitud<br>positiva",PAL[3])}
    {card_kpi(pct(dist(datos[2]["rows"],idx[2]["ns"]).get("Alto",0),datos[2]["n"]),"% Riesgo no show<br>Alto","#ef4444")}
    {card_kpi(pct(dist(datos[2]["rows"],idx[2]["br"]).get("Alta",0),datos[2]["n"]),"% Brecha precio<br>Alta","#f59e0b")}
  </div>
  <div class="g2">
    <div class="card"><h3>Actitud Inicial</h3><canvas id="ch_c2_act"></canvas></div>
    <div class="card"><h3>Brecha Precio-Expectativa</h3><canvas id="ch_c2_br"></canvas></div>
  </div>
  <div class="g2">
    <div class="card"><h3>Riesgo de No Show</h3><canvas id="ch_c2_ns"></canvas></div>
    <div class="card"><h3>Riesgo Comercial Identificado</h3><canvas id="ch_c2_rc"></canvas></div>
  </div>
  <div class="card">
    <h3>Análisis — Fricción Comercial y Señales de Riesgo</h3>
    <div class="synthesis">{sint.get(2,"")}</div>
  </div>'''}
</div>

<!-- ── CAPA 3 ── -->
<div id="c3" class="sec">
  <div class="capa-header" style="background:{PAL_LIGHT[3]};color:{PAL_TEXT[3]}">
    <div class="num">3</div>
    <div class="info">
      <h2>{ICON[3]} Capa 3 — Compra exitosa (ABR-JUN 2025 hasta el 4 de junio)</h2>
      <p>Casos de éxito · Base para el proceso de validación de alto interés · 100% con cita confirmada</p>
    </div>
  </div>
  {pending_section(3,'Analisis_Llamadas_Final.xlsx','generar_panel_analisis.py','llamadas') if 3 not in datos else f'''
  <div class="kpi-row">
    {card_kpi(datos[3]["n"],"Conversiones<br>exitosas",PAL[3])}
    {card_kpi(gp[3][0],"% Actitud<br>positiva","#22c55e")}
    {card_kpi(pct(dists[3]["pr"].get("Sí",0),datos[3]["n"]),"% Preguntaron<br>precio","#f59e0b")}
    {card_kpi(pct(dists[3]["fr"].get("No le gusta",0),datos[3]["n"]),"% Fricción<br>con precio","#ef4444")}
    {card_kpi(pct(dists[3]["co"].get("Sí",0),datos[3]["n"]),"% Cotizaron<br>en otro lado","#8b5cf6")}
    {card_kpi(pct(sum(v for k,v in dists[3]["pd"].items() if k.startswith("Sí")),datos[3]["n"]),"% Con prenda<br>o crédito","#06b6d4")}
  </div>
  <div class="g3">
    <div class="card"><h3>Actitud Inicial</h3><canvas id="ch_c3_act"></canvas></div>
    <div class="card"><h3>Motivo de Venta</h3><canvas id="ch_c3_mot"></canvas></div>
    <div class="card"><h3>Fricción en Precio</h3><canvas id="ch_c3_fr"></canvas></div>
  </div>
  <div class="g3">
    <div class="card"><h3>¿Preguntó Precio?</h3><canvas id="ch_c3_pr"></canvas></div>
    <div class="card"><h3>¿Ya Cotizó en Otro Lado?</h3><canvas id="ch_c3_co"></canvas></div>
    <div class="card"><h3>Prenda / Financiera</h3><canvas id="ch_c3_pd"></canvas></div>
  </div>
  <div class="card">
    <h3>Factores de Conversión — Análisis de Casos Exitosos</h3>
    <div class="synthesis">{sint.get(3,"")}</div>
  </div>'''}
</div>

<!-- ── COMPARACIÓN ── -->
<div id="comp" class="sec">
  {"" if len(capas_ok)>=2 else f'''
  <div class="insight" style="background:#f0f9ff;border-color:#38bdf8">
    <h3 style="color:#0369a1">Comparación disponible con 2 o más capas cargadas</h3>
    <p>Actualmente solo está cargada la capa {capas_ok[0] if capas_ok else "—"}.
    Carga los MP3 de al menos otra capa y regenera el panel para ver la comparación.</p>
  </div>'''}
  {"" if len(capas_ok)<2 else f'''
  <div class="card">
    <h3>Síntesis Comparativa — ¿Qué diferencia a cada tipo de cliente?</h3>
    <div class="synthesis">{sint.get("comp","")}</div>
  </div>
  <div class="g2">
    <div class="card"><h3>Actitud Inicial por Capa</h3><canvas id="ch_act_comp"></canvas></div>
    <div class="card"><h3>Pregunta de Precio por Capa</h3><canvas id="ch_precio_comp"></canvas></div>
  </div>
  <div class="card">
    <h3>Fricción en Precio por Capa</h3>
    <canvas id="ch_fr_comp" style="max-height:220px"></canvas>
  </div>'''}
  <div class="insight" style="background:#f8fafc;border-color:#cbd5e1;margin-top:16px">
    <h3 style="color:#475569">Marco de validación de alto interés</h3>
    <p>
      <strong>Protocolo de acción rápida</strong> — cuando el cliente muestra señales de alto interés en los primeros 60 segundos:<br><br>
      <strong style="color:{PAL[1]}">Capa 1</strong> previene: no perder al cliente por falta de manejo de objeciones.<br>
      <strong style="color:{PAL[2]}">Capa 2</strong> previene: no perder la cita por brecha de precio o expectativa no gestionada.<br>
      <strong style="color:{PAL[3]}">Capa 3</strong> define: qué perfil y qué conducta predice el cierre, y hay que replicarla.
    </p>
  </div>
</div>

<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script>
function showTab(id,el){{
  document.querySelectorAll('.sec').forEach(s=>s.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('active'));
  document.getElementById(id).classList.add('active');
  el.classList.add('active');
}}
window.addEventListener('load',()=>{{
  {charts_js}
}});
</script>
</body></html>"""

    HTML_SAL.write_text(html, encoding="utf-8")
    print(f"Panel guardado: {HTML_SAL}")
    print(f"Tamaño: {HTML_SAL.stat().st_size // 1024} KB")


if __name__ == "__main__":
    main()
