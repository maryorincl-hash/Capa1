import openpyxl
wb = openpyxl.load_workbook(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos\Analisis_Llamadas_Limpio.xlsx")
ws = wb.active
headers = [cell.value for cell in ws[1]]

print("COLUMNAS:")
for i, h in enumerate(headers, 1):
    print(f"  {i:2}. {h}")

print("\nEJEMPLO FILA COMPLETA (fila 2):")
row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
for i in range(15):
    v = str(row[i])[:80] if row[i] else "VACIO"
    print(f"  {i+1:2}. {headers[i]}: {v}")

print("\nEJEMPLO FILA 3:")
row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
for i in range(15):
    v = str(row[i])[:80] if row[i] else "VACIO"
    print(f"  {i+1:2}. {headers[i]}: {v}")

print("\nFILA 19 (vacia):")
row = list(ws.iter_rows(min_row=19, max_row=19, values_only=True))[0]
for i in range(16):
    v = str(row[i])[:80] if row[i] else "VACIO"
    print(f"  {i+1:2}. {headers[i]}: {v}")
