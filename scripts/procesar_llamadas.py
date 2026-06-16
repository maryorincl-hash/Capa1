#!/usr/bin/env python3
"""
Procesador de llamadas - Leads Compra
Transcribe MP3s con Whisper y analiza con Claude API.
Salida: Excel enriquecido en carpeta datos/
"""

import os
import json
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import anthropic
import whisper

# ============================================================
# CONFIGURACIÓN
# ============================================================
LLAMADAS_DIR  = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\llamadas")
DATOS_DIR     = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos")
EXCEL_ENTRADA = LLAMADAS_DIR / "Auditoría Completa de Contact Center - Reporte Consolidado V6.xlsx"
EXCEL_SALIDA  = DATOS_DIR / "Analisis_Llamadas_Completo.xlsx"
PROGRESO_FILE = DATOS_DIR / ".progreso.json"   # checkpoint para resumir si se corta

MODELO_WHISPER = "small"          # base=rápido, small=preciso, medium=más preciso
MODELO_CLAUDE  = "claude-haiku-4-5"  # rápido y económico para extracción estructurada

# ============================================================
# COLUMNAS NUEVAS
# ============================================================
COLUMNAS_NUEVAS = [
    "Transcripción",
    "¿Preguntó por precio?",       # Sí / No
    "Precio mencionado ($)",        # monto o "No mencionado"
    "Fricción en el precio",        # Acepta / No le gusta / N/A
    "Comentario del cliente sobre precio",
    "Motivo de venta",
    "¿Ya cotizó en otro lado?",    # Sí / No / No menciona
    "Señales de decisión de compra",
    "Sugerencias del cliente",
    "Ficha Inspector",              # resumen ejecutivo para el inspector
]

# ============================================================
# PROMPT DE ANÁLISIS
# ============================================================
PROMPT = """Eres un analista experto en contact center de compra de vehículos usados.
Esta llamada terminó en UNA COMPRA EXITOSA: el cliente agendó una cita de inspección.
Tu tarea es extraer factores que determinaron esa compra desde el primer contacto.

TRANSCRIPCIÓN:
{transcripcion}

Responde ÚNICAMENTE con JSON válido, sin texto adicional:
{{
    "pregunto_precio": "Sí" o "No",
    "precio_mencionado": "monto en $ o 'No mencionado'",
    "friccion_precio": "Acepta" o "No le gusta" o "N/A",
    "comentario_precio": "cita textual del cliente sobre el precio, o 'N/A'",
    "motivo_venta": "una de estas opciones: Necesita liquidez / Cambio de auto / Urgencia alta / Sin urgencia / No menciona",
    "cotizo_otro_lado": "Sí" o "No" o "No menciona",
    "senales_compra": "factores concretos que determinaron la decisión: actitud, preguntas, urgencia, confianza generada (máx 120 palabras)",
    "sugerencias_cliente": "qué pide o propone el cliente para mejorar el proceso, o 'Ninguna'",
    "ficha_inspector": "Resumen para el inspector: ACTITUD del cliente al llegar, URGENCIA de venta, PRECIO esperado, ESTADO del auto mencionado, ADVERTENCIAS o detalles especiales para la cita. Máx 150 palabras."
}}"""


# ============================================================
# FUNCIONES
# ============================================================

def cargar_datos_existentes():
    """Lee el Excel de entrada y retorna dict {filename: {col: valor}}."""
    wb = openpyxl.load_workbook(EXCEL_ENTRADA)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    datos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            fila = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
            datos[row[0]] = fila
    wb.close()
    return datos, [h for h in headers if h]


def cargar_progreso():
    """Carga el archivo de progreso (archivos ya procesados)."""
    if PROGRESO_FILE.exists():
        with open(PROGRESO_FILE, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def guardar_progreso(procesados):
    """Guarda el conjunto de archivos ya procesados."""
    DATOS_DIR.mkdir(parents=True, exist_ok=True)
    with open(PROGRESO_FILE, "w", encoding="utf-8") as f:
        json.dump(list(procesados), f)


def inicializar_excel(headers_originales):
    """Crea el workbook de salida si no existe, o lo abre si ya hay progreso."""
    todos_headers = headers_originales + COLUMNAS_NUEVAS

    if EXCEL_SALIDA.exists():
        wb = openpyxl.load_workbook(EXCEL_SALIDA)
        ws = wb.active
        return wb, ws, todos_headers

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Análisis Completo"
    ws.freeze_panes = "A2"

    for i, h in enumerate(todos_headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        es_nueva = h in COLUMNAS_NUEVAS
        cell.fill = PatternFill("solid", fgColor="2E75B6" if es_nueva else "1F4E79")
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    wb.save(EXCEL_SALIDA)
    return wb, ws, todos_headers


def agregar_fila(ws, fila_num, todos_headers, datos_orig, analisis):
    """Escribe una fila en el Excel."""
    for j, header in enumerate(todos_headers, 1):
        if header in analisis:
            valor = analisis[header]
        else:
            valor = datos_orig.get(header, "")
        cell = ws.cell(row=fila_num, column=j, value=valor)
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def transcribir(modelo_whisper, ruta_mp3):
    """Transcribe el MP3 con Whisper en español."""
    try:
        result = modelo_whisper.transcribe(
            str(ruta_mp3), language="es", task="transcribe", fp16=False
        )
        return result["text"].strip()
    except Exception as e:
        return f"[ERROR TRANSCRIPCIÓN: {e}]"


def analizar(client, transcripcion):
    """Envía la transcripción a Claude y retorna dict con columnas nuevas."""
    if not transcripcion or transcripcion.startswith("[ERROR"):
        return {c: "Sin transcripción" for c in COLUMNAS_NUEVAS}

    # Limitar largo para no exceder tokens
    texto = transcripcion[:9000] + ("...[cortado]" if len(transcripcion) > 9000 else "")

    try:
        response = client.messages.create(
            model=MODELO_CLAUDE,
            max_tokens=1200,
            messages=[{"role": "user", "content": PROMPT.format(transcripcion=texto)}]
        )
        raw = response.content[0].text.strip()
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        data = json.loads(raw)

        return {
            "Transcripción":                       transcripcion,
            "¿Preguntó por precio?":               data.get("pregunto_precio", "N/A"),
            "Precio mencionado ($)":               data.get("precio_mencionado", "N/A"),
            "Fricción en el precio":               data.get("friccion_precio", "N/A"),
            "Comentario del cliente sobre precio": data.get("comentario_precio", "N/A"),
            "Motivo de venta":                     data.get("motivo_venta", "N/A"),
            "¿Ya cotizó en otro lado?":            data.get("cotizo_otro_lado", "N/A"),
            "Señales de decisión de compra":       data.get("senales_compra", "N/A"),
            "Sugerencias del cliente":             data.get("sugerencias_cliente", "N/A"),
            "Ficha Inspector":                     data.get("ficha_inspector", "N/A"),
        }

    except json.JSONDecodeError:
        raw_corto = raw[:300] if "raw" in dir() else "respuesta vacía"
        return {c: f"Error JSON: {raw_corto}" for c in COLUMNAS_NUEVAS}
    except Exception as e:
        return {c: f"Error API: {str(e)[:100]}" for c in COLUMNAS_NUEVAS}


# ============================================================
# MAIN
# ============================================================

def main():
    # Verificar API key
    if not os.environ.get("ANTHROPIC_API_KEY"):
        print("ERROR: Falta ANTHROPIC_API_KEY.")
        print("Ejecuta en PowerShell: $env:ANTHROPIC_API_KEY = 'sk-ant-...'")
        return

    DATOS_DIR.mkdir(parents=True, exist_ok=True)

    print("=" * 55)
    print("  PROCESADOR DE LLAMADAS - LEADS COMPRA")
    print("=" * 55)

    # 1. Cargar Excel de entrada
    print("\n[1] Cargando Excel de entrada...")
    datos_existentes, headers_orig = cargar_datos_existentes()
    print(f"    {len(datos_existentes)} filas encontradas en el Excel base.")

    # 2. Listar MP3s
    mp3s = sorted(LLAMADAS_DIR.glob("*.mp3"))
    print(f"\n[2] {len(mp3s)} archivos MP3 encontrados.")
    if not mp3s:
        print("    No hay MP3s en la carpeta. Verifica la ruta.")
        return

    # 3. Progreso previo (para resumir si se cortó)
    procesados = cargar_progreso()
    pendientes = [p for p in mp3s if p.name not in procesados]
    if procesados:
        print(f"    Retomando: {len(procesados)} ya procesados, {len(pendientes)} pendientes.")

    # 4. Cargar modelos
    print(f"\n[3] Cargando Whisper '{MODELO_WHISPER}'...")
    modelo_whisper = whisper.load_model(MODELO_WHISPER)
    print(f"    Whisper listo. Claude: {MODELO_CLAUDE}")
    client = anthropic.Anthropic()

    # 5. Inicializar Excel de salida
    wb, ws, todos_headers = inicializar_excel(headers_orig)
    fila_num = ws.max_row + 1  # continuar desde donde quedó

    # 6. Procesar
    print(f"\n[4] Procesando {len(pendientes)} llamadas...\n")
    errores = []

    for i, ruta in enumerate(pendientes, 1):
        nombre = ruta.name
        print(f"  [{i}/{len(pendientes)}] {nombre}")

        datos_orig = datos_existentes.get(nombre, {})

        print(f"         Transcribiendo...", end=" ", flush=True)
        transcripcion = transcribir(modelo_whisper, ruta)
        print(f"{len(transcripcion)} chars")

        print(f"         Analizando con Claude...", end=" ", flush=True)
        analisis = analizar(client, transcripcion)
        print("OK" if not analisis.get("Transcripción", "").startswith("Error") else "WARN")

        agregar_fila(ws, fila_num, todos_headers, datos_orig, analisis)
        fila_num += 1

        # Marcar como procesado
        procesados.add(nombre)

        # Checkpoint cada 10 archivos
        if i % 10 == 0 or i == len(pendientes):
            wb.save(EXCEL_SALIDA)
            guardar_progreso(procesados)
            print(f"         [Checkpoint] {len(procesados)} / {len(mp3s)} guardados -> {EXCEL_SALIDA.name}")

    # 7. Ajustar anchos de columna
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value).split("\n")[0]), 60))
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

    wb.save(EXCEL_SALIDA)

    # Limpiar archivo de progreso al terminar todo
    if len(procesados) == len(mp3s):
        PROGRESO_FILE.unlink(missing_ok=True)

    print("\n" + "=" * 55)
    print(f"  COMPLETADO")
    print(f"  Excel: {EXCEL_SALIDA}")
    print(f"  Total filas: {fila_num - 2}")
    if errores:
        print(f"  Advertencias: {len(errores)} archivos con errores")
    print("=" * 55)


if __name__ == "__main__":
    main()
