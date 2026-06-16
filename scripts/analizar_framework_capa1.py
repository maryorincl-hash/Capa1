#!/usr/bin/env python3
# CAPA 1 - Re-analisis con Framework Comercial (Analisis.docx)
# Lee las transcripciones ya guardadas en .progreso_capa1.json y aplica
# el framework de 3 pilares + 5 tipificaciones de fracaso.
# NO requiere Whisper ni MP3s. Solo necesita la API key.
#
# Antes de correr:
#   $env:ANTHROPIC_API_KEY = (Get-ItemProperty "HKCU:\Environment" -Name "ANTHROPIC_API_KEY").ANTHROPIC_API_KEY
#   cd C:\Users\maryorin.vivas\Proyectos\LeadsCompra
#   python scripts\analizar_framework_capa1.py

import json, re
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import anthropic

DATOS_DIR      = Path(r"C:\Users\maryorin.vivas\Proyectos\LeadsCompra\datos")
PROGRESO_BASE  = DATOS_DIR / ".progreso_capa1.json"
PROGRESO_FILE  = DATOS_DIR / ".progreso_framework_capa1.json"
EXCEL_SALIDA   = DATOS_DIR / "Analisis_Capa1_Framework.xlsx"
MODELO_CLAUDE  = "claude-haiku-4-5"

# ── colores de sección ────────────────────────────────────────────────────────
COLOR_ID       = "1E3A5F"  # azul oscuro — datos básicos
COLOR_A        = "15803D"  # verde — Pilar A: Apertura
COLOR_B        = "1D4ED8"  # azul — Pilar B: Escucha activa
COLOR_C        = "7E22CE"  # morado — Pilar C: Cierre
COLOR_FRAC     = "DC2626"  # rojo — Tipificación del fracaso
COLOR_COACHING = "92400E"  # café — Coaching

# ── columnas ─────────────────────────────────────────────────────────────────
COLUMNAS = [
    # Identificación
    ("Nombre de Archivo",    COLOR_ID),
    ("Cliente",              COLOR_ID),
    ("Tiempo de Conversación", COLOR_ID),
    ("Vehículo / Año",       COLOR_ID),
    ("Patente",              COLOR_ID),
    ("Tipo de Precio",       COLOR_ID),
    # Condiciones del auto
    ("Licencia Clase B",     COLOR_ID),
    ("Sin Encargo Robo",     COLOR_ID),
    ("Sin Remate",           COLOR_ID),
    (">30 Días Titular",     COLOR_ID),
    ("Prenda / Financiera",  COLOR_ID),
    # Pilar A — Apertura
    ("A1 · Velocidad de Contexto",  COLOR_A),
    ("A2 · Gestión del Precio",     COLOR_A),
    ("A3 · Tono Inicial del Agente", COLOR_A),
    ("A · Calificación Apertura",   COLOR_A),
    # Pilar B — Escucha activa
    ("B1 · Preguntas de Descubrimiento",  COLOR_B),
    ("B2 · Motivación Detectada",        COLOR_B),
    ("B3 · Señales de Interés del Cliente", COLOR_B),
    ("B · Calificación Escucha",         COLOR_B),
    # Pilar C — Cierre
    ("C1 · Cita Vendida Como",           COLOR_C),
    ("C2 · Manejo Primera Negativa",     COLOR_C),
    ("C3 · Segundo Cierre Intentado",    COLOR_C),
    ("C4 · Alternativas de Agenda Ofrecidas", COLOR_C),
    ("C · Calificación Cierre",          COLOR_C),
    # Tipificación del fracaso
    ("Tipificación de Fracaso",          COLOR_FRAC),
    ("Detalle del Fracaso",              COLOR_FRAC),
    ("Objeción Principal (textual)",     COLOR_FRAC),
    ("Momento del Primer Desinterés",    COLOR_FRAC),
    ("Actitud / Perfil del Cliente",     COLOR_FRAC),
    # Coaching
    ("Recomendación de Coaching",        COLOR_COACHING),
    # Transcripción
    ("Transcripción",                    "475569"),
]

NOMBRES_COL = [c[0] for c in COLUMNAS]

# ── prompt ────────────────────────────────────────────────────────────────────
PROMPT = """Eres un auditor senior de contact center de compra de autos usados en Chile (Americar / clicar.cl).
Esta llamada TERMINÓ SIN AGENDAR CITA: el cliente atendió pero no quiso reservar inspección.
Evalúa la gestión usando el Framework Comercial de 3 Pilares.

TRANSCRIPCION:
{transcripcion}

FRAMEWORK DE EVALUACION:

PILAR A — DISCURSO DE ENTRADA (Apertura y Validacion):
  A1 Velocidad de Contexto: ¿El agente conectó de inmediato con la acción del cliente?
     Ejemplo correcto: "Hola [Nombre], veo que acabas de tasar tu [Marca] año [Año] en nuestra web..."
  A2 Gestión del Precio: ¿El precio fue automatico (web) o manual? Si fue manual, ¿vendió la llamada como "solución exclusiva"?
  A3 Tono Inicial: ¿Fue de asesoría experta, o sonó a cobranza/servicio masivo?

PILAR B — ESCUCHA ACTIVA Y DIAGNÓSTICO:
  B1 Preguntas de Descubrimiento: ¿El agente averiguó el dolor/motivación? (¿Quiere renovar?, ¿necesita dinero?, ¿solo sondea?)
  B2 Captura de Señales de Interés: Si el cliente preguntó sobre proceso/pago/tiempo → hay interés. ¿El agente lo aprovechó?

PILAR C — CONVENCIMIENTO Y CIERRE:
  C1 Cita vendida como Beneficio vs Trámite: Beneficio = "paso definitivo para garantizar ese precio, sin compromiso, 20 minutos, seguridad total"
  C2 Manejo Primera Negativa: ¿El agente rebatió el primer "no" con argumentos de oportunidad, o se rindió?
  C3 Segundo Cierre: ¿Intentó al menos un segundo cierre después de la primera negativa?
  C4 Alternativas de Agenda: ¿Ofreció horarios alternativos, sucursales distintas, o llamada posterior?

TIPIFICACIONES DE FRACASO (elige la principal):
  1. "Brecha de Precio" — El cliente dice que el precio está muy por debajo de lo que espera.
  2. "Falta de Propiedad/Decisión" — El auto es de un familiar, está en prenda, o no es el tomador de decisiones.
  3. "Fricción Logística" — Interés pero sucursal lejos, horarios no le acomodan, no puede mover el auto.
  4. "Timing — Solo Evaluando" — Planea vender en 3-6 meses, solo estaba indexando el valor de su patrimonio.
  5. "Falla del Agente" — El cliente estaba abierto, pero el agente sonó robótico, no manejó la objeción, se despidió rápido o no ofreció alternativas.

PERFILES DE CLIENTE (elige el más cercano):
  - Pragmático: directo, quiere datos concretos, poco tiempo
  - Emocional: habla de necesidades personales, responde a la confianza
  - Analítico: hace preguntas, quiere entender el proceso completo
  - Desconfiado: escéptico, resistente, necesita validación
  - Timing (solo evaluando): no tiene urgencia real, solo consultando precio
  - Frustrado/Precio: molesto porque el precio no cumple su expectativa

Responde UNICAMENTE con JSON valido, sin texto adicional:
{{
    "cliente": "nombre o Desconocido",
    "tiempo_conversacion": "MM:SS estimado",
    "vehiculo_anio": "Marca Modelo (AAAA) o No mencionado",
    "patente": "patente chilena o No mencionada",
    "tipo_precio": "Automático web / Manual / No mencionado",
    "licencia_clase_b": "Si / No / No menciona",
    "sin_encargo_robo": "Si / No / No menciona",
    "sin_remate": "Si / No / No menciona",
    "dias_titular": "Si / No / No menciona",
    "prenda_financiera": "Si (institucion si se nombra) / No / No menciona",

    "a1_velocidad_contexto": "Sí conectó con la tasación / Genérico (no mencionó el auto) / No se identificó correctamente",
    "a2_gestion_precio": "descripcion breve de como manejó el precio. Max 60 palabras.",
    "a3_tono_inicial": "Experto/Asesor / Neutro / Robótico/Monótono / Agresivo/Presión",
    "a_calificacion": "Excelente / Buena / Regular / Deficiente",

    "b1_preguntas_descubrimiento": "Sí hizo preguntas / Parcialmente / No preguntó nada",
    "b2_motivacion_detectada": "descripcion de la motivación del cliente si se reveló. 'No detectada' si el agente no preguntó. Max 60 palabras.",
    "b3_senales_interes": "descripcion de señales de interés del cliente y si el agente las aprovechó. 'Sin señales' si el cliente nunca mostró interés. Max 60 palabras.",
    "b_calificacion": "Excelente / Buena / Regular / Deficiente",

    "c1_cita_como": "Beneficio (sin compromiso, rápido) / Trámite (obligatorio/proceso) / No intentó vender la cita",
    "c2_primera_negativa": "Rebatió con argumento / Aceptó el no sin insistir / No hubo negativa explícita",
    "c3_segundo_cierre": "Sí intentó / No intentó / No aplicaba (no hubo primer cierre)",
    "c4_alternativas": "Ofreció alternativas de horario/sucursal / Solo una opción / No ofreció nada",
    "c_calificacion": "Excelente / Buena / Regular / Deficiente",

    "tipificacion_fracaso": "Brecha de Precio / Falta de Propiedad-Decisión / Fricción Logística / Timing — Solo Evaluando / Falla del Agente",
    "detalle_fracaso": "explicacion especifica de por qué no se agendó, citando el momento o frase clave. Max 100 palabras.",
    "objecion_principal": "objeción o frase del cliente que determinó el fracaso. Citar textual si es posible.",
    "momento_primer_desinteres": "descripcion del momento de la llamada en que el cliente mostró el primer desinterés. Max 60 palabras.",
    "perfil_cliente": "Pragmático / Emocional / Analítico / Desconfiado / Timing (solo evaluando) / Frustrado-Precio",

    "recomendacion_coaching": "recomendacion especifica y concreta para el agente basada en esta llamada. Qué debería haber dicho o hecho diferente. Max 100 palabras."
}}"""


# ── utilidades ────────────────────────────────────────────────────────────────

def cargar_json(path):
    if path.exists():
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return {}

def guardar_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def analizar(client, transcripcion, archivo):
    texto = transcripcion[:9000]
    try:
        r = client.messages.create(
            model=MODELO_CLAUDE, max_tokens=1600,
            messages=[{"role": "user", "content": PROMPT.format(transcripcion=texto)}]
        )
        raw = r.content[0].text.strip()
        raw = re.sub(r"^```json\s*", "", raw)
        raw = re.sub(r"\s*```$", "", raw)
        return json.loads(raw)
    except Exception as e:
        print(f"  ERROR {archivo}: {e}")
        return None

def crear_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Framework Capa 1"
    for j, (nombre, color) in enumerate(COLUMNAS, 1):
        c = ws.cell(row=1, column=j, value=nombre)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"
    return wb, ws

def escribir_fila(ws, n, d):
    for j, (nombre, _) in enumerate(COLUMNAS, 1):
        c = ws.cell(row=n, column=j, value=d.get(nombre, ""))
        c.alignment = Alignment(wrap_text=True, vertical="top")

def ajustar_anchos(ws):
    ANCHOS = {
        "Transcripción": 55, "Recomendación de Coaching": 42,
        "Detalle del Fracaso": 42, "B2 · Motivación Detectada": 38,
        "B3 · Señales de Interés del Cliente": 38, "Momento del Primer Desinterés": 38,
        "A2 · Gestión del Precio": 35, "Objeción Principal (textual)": 35,
        "C2 · Manejo Primera Negativa": 35,
    }
    for col in ws.columns:
        h = col[0].value or ""
        max_len = max((min(len(str(c.value or "").split("\n")[0]), 55) for c in col[1:]), default=0)
        ws.column_dimensions[col[0].column_letter].width = ANCHOS.get(h, max(13, max_len + 2))

# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 65)
    print("  CAPA 1 — RE-ANÁLISIS CON FRAMEWORK COMERCIAL")
    print("=" * 65)

    base = cargar_json(PROGRESO_BASE)
    if not base:
        print(f"\n  No se encontró {PROGRESO_BASE}")
        print("  Primero corre procesar_capa1.py para generar las transcripciones.")
        return

    # Solo los que tienen transcripción válida
    transcripciones = {
        k: v.get("Transcripcion", v.get("transcripcion", ""))
        for k, v in base.items()
        if v.get("Transcripcion") and not str(v.get("Transcripcion","")).startswith("Error")
    }
    print(f"\n  Transcripciones disponibles: {len(transcripciones)}")

    progreso = cargar_json(PROGRESO_FILE)
    ya_hechos = set(progreso.keys())
    pendientes = [(k, t) for k, t in transcripciones.items() if k not in ya_hechos]

    if not pendientes:
        print("  Todos ya analizados. Regenerando Excel...")
    else:
        print(f"  Pendientes: {len(pendientes)}  (ya hechos: {len(ya_hechos)})")

    client = anthropic.Anthropic()

    for idx, (archivo, transcripcion) in enumerate(pendientes, 1):
        print(f"  [{idx}/{len(pendientes)}] {archivo}", end=" -> ", flush=True)
        datos = analizar(client, transcripcion, archivo)
        if datos is None:
            print("FALLO")
            continue

        progreso[archivo] = {
            "Nombre de Archivo":                archivo,
            "Cliente":                          datos.get("cliente", "Desconocido"),
            "Tiempo de Conversación":           datos.get("tiempo_conversacion", ""),
            "Vehículo / Año":                   datos.get("vehiculo_anio", "No mencionado"),
            "Patente":                          datos.get("patente", "No mencionada"),
            "Tipo de Precio":                   datos.get("tipo_precio", "No mencionado"),
            "Licencia Clase B":                 datos.get("licencia_clase_b", "No menciona"),
            "Sin Encargo Robo":                 datos.get("sin_encargo_robo", "No menciona"),
            "Sin Remate":                       datos.get("sin_remate", "No menciona"),
            ">30 Días Titular":                 datos.get("dias_titular", "No menciona"),
            "Prenda / Financiera":              datos.get("prenda_financiera", "No menciona"),
            "A1 · Velocidad de Contexto":       datos.get("a1_velocidad_contexto", ""),
            "A2 · Gestión del Precio":          datos.get("a2_gestion_precio", ""),
            "A3 · Tono Inicial del Agente":     datos.get("a3_tono_inicial", ""),
            "A · Calificación Apertura":        datos.get("a_calificacion", ""),
            "B1 · Preguntas de Descubrimiento": datos.get("b1_preguntas_descubrimiento", ""),
            "B2 · Motivación Detectada":        datos.get("b2_motivacion_detectada", ""),
            "B3 · Señales de Interés del Cliente": datos.get("b3_senales_interes", ""),
            "B · Calificación Escucha":         datos.get("b_calificacion", ""),
            "C1 · Cita Vendida Como":           datos.get("c1_cita_como", ""),
            "C2 · Manejo Primera Negativa":     datos.get("c2_primera_negativa", ""),
            "C3 · Segundo Cierre Intentado":    datos.get("c3_segundo_cierre", ""),
            "C4 · Alternativas de Agenda Ofrecidas": datos.get("c4_alternativas", ""),
            "C · Calificación Cierre":          datos.get("c_calificacion", ""),
            "Tipificación de Fracaso":          datos.get("tipificacion_fracaso", ""),
            "Detalle del Fracaso":              datos.get("detalle_fracaso", ""),
            "Objeción Principal (textual)":     datos.get("objecion_principal", ""),
            "Momento del Primer Desinterés":    datos.get("momento_primer_desinteres", ""),
            "Actitud / Perfil del Cliente":     datos.get("perfil_cliente", ""),
            "Recomendación de Coaching":        datos.get("recomendacion_coaching", ""),
            "Transcripción":                    transcripcion,
        }
        print("OK")

        if idx % 10 == 0:
            guardar_json(PROGRESO_FILE, progreso)
            print(f"    [Checkpoint {idx}/{len(pendientes)}]")

    guardar_json(PROGRESO_FILE, progreso)

    print("\n  Generando Excel...")
    wb, ws = crear_excel()
    for n, (archivo, datos) in enumerate(progreso.items(), 2):
        escribir_fila(ws, n, datos)
    ajustar_anchos(ws)
    wb.save(EXCEL_SALIDA)

    total = len(progreso)
    print(f"\n{'=' * 65}")
    print(f"  COMPLETADO — {total} llamadas analizadas con framework")
    print(f"  Excel: {EXCEL_SALIDA}")
    print(f"{'=' * 65}")


if __name__ == "__main__":
    main()
